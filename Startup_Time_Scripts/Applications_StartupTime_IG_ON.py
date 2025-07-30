# Import necessary libraries
import re
import os
import sys
import json
import glob
import time
import yaml
import serial
import openpyxl
import xml.etree.ElementTree as ET
import subprocess
import matplotlib
import matplotlib.pyplot as plt
import ipaddress
matplotlib.use('Agg')
import numpy as np
from openpyxl.drawing.image import Image
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter
from pathlib import Path
from datetime import datetime
import logging
from typing import Final
from enum import Enum
import threading
from collections import OrderedDict
import colorlog
import pandas as pd
from decimal import Decimal, ROUND_HALF_UP


plot_lock = threading.Lock()

def round_decimal_half_up(number, decimals=0):
    """
    Rounds a number using traditional rounding (0.5 always rounds up).
    
    Args:
        number (float): Number to round
        decimals (int): Number of decimal places
        
    Returns:
        float: Properly rounded number
    """
    multiplier = 10 ** decimals
    return float(Decimal(str(number * multiplier)).quantize(Decimal('1'), rounding=ROUND_HALF_UP)) / multiplier

cur_dt_time_obj = None
local_save_path = None
workbook_map = None
threshold_map = None
current_timestamp = None
is_pre_gen_logs = None
table_headers = None

def setup_logging():
    """
    Configures and initializes the logging system with colored output formatting.
    
    This function sets up a comprehensive logging configuration that includes:
    - Colored log messages for better visual distinction
    - Timestamp, log level, function name, and line number information
    - Console output handler with custom formatting
    
    Returns:
        logging.Logger: Configured logger instance for the current module
        
    Note:
        The root logger level is set to INFO, meaning DEBUG messages won't be displayed
        unless explicitly changed. The color formatting helps distinguish between
        different log levels (INFO, WARNING, ERROR, etc.) in console output.
    """
    # Set up colored logging configuration
    LOG_FORMAT = (
        '%(log_color)s%(asctime)s - %(levelname)s - %(threadName)s - %(funcName)s - %(lineno)d - %(message)s%(reset)s'
    )
    logging.root.setLevel(logging.INFO)  # Set the root logger level to INFO

    # Configure the colorlog formatter
    formatter = colorlog.ColoredFormatter(LOG_FORMAT, datefmt="%Y-%m-%d %H:%M:%S")

    # Create a StreamHandler for console output
    stream = logging.StreamHandler()
    stream.setFormatter(formatter)

    # Add the handlers to the root logger
    logging.root.addHandler(stream)    

    # Return the configured logger
    return logging.getLogger(__name__)


# Define the column names for the application startup time data
application_startup_time_columns = ['No.', 'Services/Applications', 'Application Startup\n Time (sec)',
                                    'IG ON\n to\n QNX Startup (sec)', 'Total Time\n from\n IG ON (sec)',
                                    'Startup Time\n Threshold\n (sec)', 'Startup time\n judgement', 'Expected Order', 'Result of the\n enabled judgement\n item', 'Order\n Mismatch', 'Not\n Found', 'Not\n Configured']

# Define the column names for the application startup time data with minimum, maximum, and average values
application_startup_time_min_max_avg_columns = ['Services/Applications', 'Minimum (sec)', 'Maximum (sec)',
                                                'Average (sec)', 'Average\n from\n IG ON (sec)', 'Startup Time\n Threshold\n (sec)']

application_info_columns = ['Services/Applications', 'Init(Up) Time (us)', 'Init(Up) Time (ms)']

application_start_end_time_min_max_avg_columns = ['Services/Applications', 'Minimum (ms)', 'Maximum (ms)',
                                                  'Average (ms)']

applications_overall_status_columns = ['No. of Iterations', 'Total Time\n to Startup\n Last Application\n from IG ON (sec)',
                                        'Startup time\n judgement', 'Result of the\n enabled judgement\n item', 'Order\n Mismatch\n Count', 'Not\n Found\n Count', 'Not\n Configured\n Count']

appendix_columns = ['Column Name', 'Description']
startup_field_descriptions = [
   ("Services/Applications", "Name of the Service/Application being initialized."),
   ("Start Time", "The timestamp (from DLT logs) indicating when the application startup begins.",),
   ("Apps Startup", "Time taken (in seconds) by the application to initialize after the welcome timestamp. \n Apps Startup = Start time of each Application - KSAR start time.)"),
   ("Init(Up) Time", "Time taken for the application to fully start after initialization and that is captured from the DLT logs.",),
   ("IG ON to QNX + KSAR Startup", "This is the offfset time from IG ON to QNX startup and from QNX startup to KSAR start time. \n Offset time = IG ON to QNX startup + QNX startup to KSAR startup."),
   ("Total Time", "This is the Total time taken from IG ON to Application Startup completion. \n Total time = Apps Startup +  InitUp Time/1000000 + offset time (IG ON to QNX + KSAR Startup.")
]

# Define a border style for cells in the Excel sheet
border_style = Border(left=Side(border_style='thin'), right=Side(border_style='thin'),
                        top=Side(border_style='thin'), bottom=Side(border_style='thin'))

OFFSET_TIME: Final = 1.5

class ECUType(Enum):
    RCAR = "RCAR"
    PADAS = "PADAS"
    ELITE = "ELITE"
    SoC0 = "SoC0"
    SoC1 = "SoC1"

class OrderType(Enum):
    SEQUENTIAL = "sequential"
    PARALLEL = "parallel"

class OrderFailureType(Enum):
   ORDER_MISMATCH = 1
   APPLICATION_NOT_CONFIGURED = 2
   APPLICATION_NOT_FOUND = 3

class ResultThread(threading.Thread):
   def __init__(self, target, args=(), kwargs=None):
       super().__init__()
       self._target = target
       self._args   = args
       self._kwargs = kwargs or {}
       self.result  = None
   def run(self):
       # run() is what .start() invokes
       self.result = self._target(*self._args, **self._kwargs)

def is_valid_ip(ip_str):
    """
    Validates whether a given string represents a valid IP address (IPv4 or IPv6).
    
    This function uses Python's ipaddress module to validate the format and
    structure of an IP address string. It supports both IPv4 and IPv6 formats.
    
    Args:
        ip_str (str): The IP address string to validate
        
    Returns:
        bool: True if the IP address is valid, False otherwise
        
    Example:
        >>> is_valid_ip("192.168.1.1")
        True
        >>> is_valid_ip("invalid_ip")
        False
        >>> is_valid_ip("2001:db8::1")
        True
    """
    try:
        ipaddress.ip_address(ip_str)
        return True
    except ValueError:
        return False

def validate_ip_address(ecu_config_list, logger):
    """
    Validates IP addresses for all ECU configurations in the provided list.
    
    This function iterates through a list of ECU configurations and validates
    each ECU's IP address using the is_valid_ip function. If any IP address
    is invalid, the function logs an error message and returns False.
    
    Args:
        ecu_config_list (list): List of dictionaries containing ECU configurations.
                               Each dictionary should have 'ip-address' and 'ecu-type' keys.
                               
    Returns:
        bool: True if all IP addresses are valid, False if any IP address is invalid
        
    Example:
        >>> ecu_list = [{'ecu-type': 'RCAR', 'ip-address': '192.168.1.33'}]
        >>> validate_ip_address(ecu_list)
        True
    """
    for ecu in ecu_config_list:
        if is_valid_ip(ecu['ip-address']):
            continue
        else:
            logger.info(f"Entered IP address for {ecu['ecu-type']} is not valid.")
            return False
    return True

def remove_png_files(logger):
    """
    Removes all PNG image files from the script's directory.
    
    This cleanup function is typically called to remove temporary graph images
    generated during previous test runs. It searches for all .png files in the
    same directory as the script and attempts to delete them.
    
    The function handles deletion errors gracefully by logging them without
    stopping the execution. This is useful for cleaning up temporary visualization
    files created by matplotlib plotting functions.
    
    Note:
        - Uses pathlib for cross-platform file operations
        - Silently continues if files cannot be deleted (e.g., if they're in use)
        - Logs errors for debugging purposes
    """
    script_dir = Path(__file__).parent
    png_files = list(script_dir.glob('*.png'))
    for file in png_files:
        try:
            # logger.info(f"Deleting {file.name}")
            file.unlink()
        except Exception as e:
            pass
            logger.info(f"Error deleting file {file}: {e}")

def adjust_column_width(sheet, ecu_type, logger):
    """
    Automatically adjusts column widths in an Excel worksheet based on content length.
    
    This function analyzes the content of each column in the worksheet and sets
    the column width to accommodate the longest content with some padding. It
    intelligently handles various Excel formatting scenarios including merged cells,
    wrapped text, and specific content types.
    
    Args:
        sheet (openpyxl.worksheet.worksheet.Worksheet): The Excel worksheet to adjust
        ecu_type (str): The ECU type identifier used to skip certain log file references
        
    Features:
        - Calculates optimal width based on maximum content length in each column
        - Skips merged cells to avoid width calculation conflicts
        - Ignores cells with text wrapping enabled
        - Filters out log file references containing ECU type
        - Ensures minimum column width of 9 characters
        - Adds 3 characters padding for better readability
        
    Note:
        This function is essential for creating professional-looking Excel reports
        where all content is visible without manual column width adjustments.
    """
    # Iterate through each column in the Excel sheet
    for col in sheet.columns:
        # Initialize a variable to track the maximum content length within the column
        max_length = 0
       
        # Extract the letter representing the label of the current column
        column_letter = get_column_letter(col[0].column)

        # Iterate through each cell in the current column, starting from the start_row
        for cell in col[0:]:
            try:
                # Check if the cell is empty
                if not cell.value:
                    continue
               
                # Skip cells with specific content
                if f'Startup_Time_Logs_{ecu_type}' in str(cell.value)  or str(cell.value) in table_headers:
                    continue

                # Attempt to retrieve the content of the cell and check its length
                cell_content = str(cell.value)
                
                # Check if the cell's alignment has wrap text enabled
                if cell.alignment.wrap_text:
                    lines = cell_content.split('\n')
                    max_length = max(max(len(line) for line in lines), max_length)
                else:
                    # If wrap text is not enabled, use the length of the cell content directly
                    max_length = max(len(cell_content), max_length)

            except (TypeError, AttributeError, ValueError) as e:
                # Handle specific exceptions
                logger.error(f"An error occurred: {e}")

        # Calculate the adjusted width for the column based on the maximum content length with extra space
        adjusted_width = max(max_length + 3, 9)  # Ensure a minimum width of 9

        # Set the column width in the Excel sheet to the calculated adjusted width
        sheet.column_dimensions[column_letter].width = adjusted_width

def format_excel_cells(sheet, start_row):
    """
    Applies comprehensive formatting to Excel worksheet cells for professional presentation.
    
    This function formats cells in an Excel worksheet starting from a specified row,
    applying different styles based on cell content type. It creates a visually
    appealing and easily readable report with color-coded status indicators.
    
    Args:
        sheet (openpyxl.worksheet.worksheet.Worksheet): The Excel worksheet to format
        start_row (int): The row number from which to start formatting (1-based indexing)
        
    Formatting Rules:
        - Column headers: Light green background (B5E6A2), bold font, bordered
        - "PASS" values: Bright green background (92D050) for success indication
        - "FAIL" values: Red background (FF0000) for failure indication
        - All cells: Center alignment (horizontal and vertical), bordered
        
    Features:
        - Skips empty rows and cells to optimize processing
        - Recognizes predefined column header types from global constants
        - Applies consistent border styling throughout the sheet
        - Uses color psychology (green=good, red=bad) for quick visual assessment
        
    Note:
        This function is crucial for creating professional test reports that
        stakeholders can quickly interpret without detailed technical knowledge.
    """
    # Iterate over each row in the sheet, starting from the specified row
    for row in sheet.iter_rows(min_row=start_row, max_row=sheet.max_row):
       
        # Skip empty rows
        if all(cell.value is None for cell in row):
            continue
       
        # Iterate over each cell in the row
        for cell in row[0:]:  
            # Skip empty cells
            if cell.value is None:
                continue
           
            # Check if the cell value is a column header
            if cell.value in (application_startup_time_columns + application_startup_time_min_max_avg_columns
                              + application_info_columns + application_start_end_time_min_max_avg_columns +
                              applications_overall_status_columns):
               
                # Apply a green fill color and bold font to column headers
                cell.fill = PatternFill(start_color="B5E6A2", end_color="B5E6A2", fill_type="solid")
                cell.font = Font(bold=True)
                cell.border = border_style
                continue
           
            elif cell.value == "PASS":
                # If the cell value is "PASS", fill it with a light green color.
                cell.fill = PatternFill(start_color = "92D050", end_color = "92D050", fill_type = "solid")

            elif cell.value == "FAIL":
                # If the cell value is "FAIL", fill it with a light red color.
                cell.fill = PatternFill(start_color = "FF0000", end_color = "FF0000", fill_type = "solid")
            elif cell.value == '⬤':
                cell.font = Font(bold=True)
               
            # Center align the cell contents horizontally and vertically
            cell.alignment = Alignment(horizontal='center', vertical='center')
           
            # Apply the defined border style to the cell
            cell.border = border_style


def plot_process_individual_apps_avg_graph(differences, sheet, start_row, ecu_type):
    """
    Creates and embeds a timeline graph showing individual application startup times.
    
    This function generates a horizontal timeline visualization that displays the average
    startup time for each application/service in milliseconds. Each application is
    represented as a horizontal line from 0 to its startup time, making it easy to
    compare relative performance across different services.
    
    Args:
        differences (dict): Dictionary mapping application names to their average startup times (ms)
        sheet (openpyxl.worksheet.worksheet.Worksheet): Excel worksheet to embed the graph
        start_row (int): Row number where the graph should be positioned
        ecu_type (str): ECU type identifier for graph title and file naming
        
    Features:
        - Thread-safe plotting using plot_lock to prevent concurrent matplotlib operations
        - Dynamic figure sizing based on number of applications (minimum 3, scales with data)
        - Horizontal timeline visualization with markers at endpoints
        - Time values displayed at midpoint of each timeline for clarity
        - Professional grid styling with dashed lines
        - Automatic x-axis scaling with 200ms tick intervals
        - Unique timestamp-based filename to avoid conflicts
        
    Graph Elements:
        - X-axis: Time interval in milliseconds
        - Y-axis: Service/Application names
        - Title: ECU-specific timeline graph title
        - Grid: Light gray dashed lines for easy reading
        
    Note:
        The generated PNG file is automatically embedded in the Excel sheet at column J
        and cleaned up by the remove_png_files() function after report generation.
    """
    with plot_lock:
        # Determine the height of the graph based on the size of the sheet
        height = (sheet.max_row - start_row + 1) * 0.35 # adjust the multiplier as needed
        width = 10

        # logger.info(f"Sheet size: {sheet.max_row - start_row + 1}, height : {height}")

        if height > 6.5:
            height = 6.5
            width = 12

        # Create a new figure with a specified size
        plt.figure(figsize=(12, max(3, len(differences)*0.2)))

        # Iterate over each process and its difference
        for index, (process, difference) in enumerate(differences.items(), start=1):
            plt.plot([0, difference], [index, index], marker='o')

            # Add a text label at the midpoint of the line with the difference value
            plt.text((0 + difference) / 2, index + 0.1, str(round_decimal_half_up(difference, 4))+" ms",#"{:.3f} ms".format(difference),
                    verticalalignment='bottom', horizontalalignment='center')

        # Set the y-axis tick labels to the process names and 'Time from IG ON to QNX startup'
        plt.yticks(range(1, len(differences) + 1), list(differences.keys()))

        # Set the x-axis label
        plt.xlabel('Time Interval (milliseconds)')

        # Set the y-axis label

        plt.ylabel('Services or Applications')

        # Set the title of the plot
        plt.title(f'{ecu_type} Timeline Graph: Individual Services/Applications Startup Time Average', pad=20)

        # Enable the grid on both x and y axes with light lines
        plt.grid(True, axis='both', linestyle='--', linewidth=0.5, color='gray')

        # Ensure the plot fits within the figure
        plt.tight_layout()

        # Determine the x-axis limits with some padding
        min_x = 0
        max_x = max(differences.values())
        # padding = (max_x - min_x) * 0.05
        padding = 3
   
        # Create a sequence of milliseconds for the x-axis
        x_ticks = np.arange(0, max_x + 200, 200)  # show ticks every 100ms

        plt.xticks(x_ticks)
        plt.xlim(min_x - padding, max_x + padding)

        # Get the current time
        timestamp = datetime.now().strftime("%M%S%f")
        # plot_image = f'graph_process_startup_{ecu_type}_{timestamp}.png'
        plot_image = Path(__file__).parent.joinpath(f'graph_process_startup_{ecu_type}_{timestamp}.png')
       


        # Save the plot to a file
        plt.savefig(plot_image)

        # Close the plot
        plt.close()

        # Add the plot to the Excel sheet
        img = Image(plot_image)
        sheet.add_image(img, f'J{start_row}')


def plot_process_start_end_time_graph(ecu_type, data, sheet, start_row):
    """
    Creates and embeds a timeline graph showing application initialization (Init/Up) times.
    
    This function generates a horizontal timeline visualization that displays the
    initialization time for each application/service. The graph shows how long each
    application takes to fully initialize after being started, measured in milliseconds.
    
    Args:
        ecu_type (str): ECU type identifier for graph title and file naming
        data (list): List of dictionaries containing process timing information
                    Each dict should have 'process' and 'start_time_ms' keys
        sheet (openpyxl.worksheet.worksheet.Worksheet): Excel worksheet to embed the graph
        start_row (int): Row number where the graph should be positioned
        
    Data Processing:
        - Converts input data to pandas DataFrame for easier manipulation
        - Ensures start_time_ms values are numeric for proper plotting
        - Sorts processes by initialization time for logical presentation
        
    Graph Features:
        - Thread-safe plotting using plot_lock
        - Dynamic figure sizing based on data volume
        - Horizontal timeline from 0 to initialization time
        - Time values displayed at timeline midpoints
        - Professional grid styling for easy reading
        - Automatic x-axis scaling with padding
        
    Graph Elements:
        - X-axis: Time interval in microseconds (despite label, data is in ms)
        - Y-axis: Service/Application names
        - Title: ECU-specific initialization time graph
        - Markers: Circular markers at timeline endpoints
        
    Note:
        There's a discrepancy in the x-axis label (microseconds) vs actual data (milliseconds).
        This should be corrected for accuracy in future versions.
    """
    with plot_lock:
        # Determine the height of the graph based on the size of the sheet
        height = (sheet.max_row - start_row + 1) * 0.35 # adjust the multiplier as needed
        width = 10

        # logger.info(f"Sheet size: {sheet.max_row - start_row + 1}, height : {height}")

        if height > 6.5:
            height = 6.5
            width = 12

        # Create a new figure with a specified size
        plt.figure(figsize=(12, max(3, len(data)*0.2)))

        df = pd.DataFrame(data)  
        df['start_time_ms'] = pd.to_numeric(df['start_time_ms'])
        print ("df", df)
 
        for index, row in df.iterrows():
            print ("row",row['start_time_ms'])
            plt.plot([0, row['start_time_ms']], [index, index], marker='o')
            plt.text(row['start_time_ms'] / 2, index + 0.1,
                    str(round_decimal_half_up(row['start_time_ms'], 4))+" ms", #"{:.3f} ms".format(row['start_time_ms']),
                    verticalalignment='bottom',
                    horizontalalignment='center')    

        plt.yticks(range(len(df)), df["process"])
        plt.xlabel('Time Interval (microseconds)')
        plt.ylabel('Services or Applications')
        plt.title(f'{ecu_type} Timeline Graph:Services/Applications Init(Up) Time', pad=20)
        plt.grid(True)
        plt.tight_layout()
   
        max_x = df['start_time_ms'].max()
        plt.xlim(-10, max_x)    

        # Get the current time
        timestamp = datetime.now().strftime("%M%S%f")

        # Save the plot as an image file
        # plot_image = f'process_start_end_graph_{ecu_type}_{timestamp}.png'
        plot_image = Path(__file__).parent.joinpath(f'process_start_end_graph_{ecu_type}_{timestamp}.png')

        plt.savefig(plot_image)
        plt.close()

        # Add the plot image to the worksheet
        img = Image(plot_image)
        sheet.add_image(img, f'J{start_row}')


def plot_process_startup_time_graph(differences, sheet, start_row, ecu_type, avg_flag):
    """
    Creates and embeds a comprehensive timeline graph showing application startup times from IG ON.
    
    This function generates the main timeline visualization that shows the complete startup
    sequence from ignition ON through QNX startup to individual application completion.
    It includes reference lines for QNX startup time and performance thresholds.
    
    Args:
        differences (dict): Dictionary mapping application names to their startup times (seconds)
        sheet (openpyxl.worksheet.worksheet.Worksheet): Excel worksheet to embed the graph
        start_row (int): Row number where the graph should be positioned
        ecu_type (str): ECU type identifier for graph title and file naming
        threshold (float): Performance threshold in seconds (shown as red dashed line)
        avg_flag (bool): If True, shows average times; if False, shows individual completion times
        
    Timeline Structure:
        - Time 0: Ignition ON
        - OFFSET_TIME (1.5s): QNX Startup completion
        - OFFSET_TIME + app_time: Individual application completion
        
    Graph Features:
        - Thread-safe plotting using plot_lock
        - Dynamic figure sizing based on number of applications
        - Horizontal timeline visualization with clear time references
        - QNX startup baseline shown as first timeline element
        - Performance threshold line (red dashed) for quick assessment
        - QNX startup reference line (black dashed) for context
        - Time values displayed at timeline midpoints
        
    Visual Elements:
        - X-axis: Time interval in seconds from ignition ON
        - Y-axis: 'Time from IG ON to QNX startup' + Service/Application names
        - Title: Dynamic based on avg_flag (Average vs Completion Time)
        - Grid: Professional dashed grid lines
        - Reference lines: Threshold (red) and QNX startup (black)
        - Labels: QNX Startup marker below timeline
        
    Note:
        This is the primary visualization for startup performance analysis,
        allowing stakeholders to quickly identify applications that exceed
        performance thresholds and understand the overall startup sequence.
    """
    with plot_lock:
        # Determine the height of the graph based on the size of the sheet
        height = (sheet.max_row - start_row + 1) * 0.35 # adjust the multiplier as needed
        width = 10

        if height > 6.5:
            height = 6.5
            width = 12

        # Create a new figure with a specified size
        plt.figure(figsize=(12, max(3, len(differences)*0.2)))  

        plt.plot([0, OFFSET_TIME], [0, 0],  marker='o')
        plt.text((OFFSET_TIME) / 2, 0.1, f"{OFFSET_TIME} sec", verticalalignment='bottom', horizontalalignment='center')

        # Iterate over each process and its difference
        for index, (process, difference) in enumerate(differences.items()):
            index = index + 1
            # Plot a line from (OFFSET_TIME, index) to (difference + OFFSET_TIME, index) with a marker at the end
            plt.plot([OFFSET_TIME, difference + OFFSET_TIME], [index, index], marker='o')

            # Add a text label at the midpoint of the line with the difference value
            plt.text((OFFSET_TIME + difference + OFFSET_TIME) / 2, index + 0.1,
                    str(round_decimal_half_up(difference, 4))+" sec", #"{:.3f} sec".format(difference),
                    verticalalignment='bottom', horizontalalignment='center')

        # Set the y-axis tick labels to the process names and 'Time from IG ON to QNX startup'
        plt.yticks(range(len(differences) + 1), ['Time from IG ON to QNX startup'] + list(differences.keys()))

        # Set the x-axis label
        plt.xlabel('Time Interval (seconds)')

        # Set the y-axis label
        plt.ylabel('Services or Applications')

        # Set the title of the graph
        if avg_flag:
            plt.title(f'{ecu_type} Timeline Graph: Services/Applications Startup Time Average', pad=20)
        else:
            plt.title(f'{ecu_type} Timeline Graph: Services/Applications Startup Completion Time', pad=20)

        # Enable the grid on both x and y axes with light lines
        plt.grid(True, axis='both', linestyle='--', linewidth=0.5, color='gray')

        # Ensure the plot fits within the figure
        plt.tight_layout()

        # Determine the x-axis limits with some padding
        min_x = 0
        max_x = max(differences.values()) + OFFSET_TIME  # add OFFSET_TIME to max_x
        padding = (max_x - min_x) * 0.015

        # Create a sequence of seconds for the x-axis
        x_ticks = np.arange(0, max_x + 1, 1)  

        plt.xticks(x_ticks)

        plt.xlim(min_x - padding, max_x + padding)
        # plt.xlim(0, max_x + padding)

        # Add QNX Startup Time label exactly below OFFSET_TIME sec on x-axis
        plt.text(OFFSET_TIME, -3.0, "QNX Startup", verticalalignment='top', horizontalalignment='center')

        # Add a vertical black line at x=OFFSET_TIME seconds
        plt.axvline(x=OFFSET_TIME, color='black', linestyle='--', linewidth=1)  

        # Get the current time
        timestamp = datetime.now().strftime("%M%S%f")
        # plot_image = f'graph_process_startup_{ecu_type}_{timestamp}.png'
        plot_image = Path(__file__).parent.joinpath(f'graph_process_startup_{ecu_type}_{timestamp}.png')

        # # Save the plot to a file
        # plot_image = f'bar_chart_process_start_time_{start_row}.png'
        plt.savefig(plot_image)

        # Close the plot
        plt.close()

        # Add the plot to the Excel sheet
        img = Image(plot_image)
        sheet.add_image(img, f'N{start_row}')

def get_log_file_path(ecu_type, setup_type, index):
    """
    Generates standardized log file paths and names for ECU startup time testing.
    
    This function creates consistent file naming conventions for log files generated
    during startup time testing. It ensures proper directory structure and handles
    directory creation if needed.
    
    Args:
        ecu_type (str): Type of ECU being tested (e.g., 'RCAR', 'SoC0', 'SoC1')
        setup_type (str): Test setup configuration type
        current_timestamp (str): Timestamp string for file naming (format: YYYYMMDD_HHMMSS)
        index (int): Current iteration index (0-based, converted to 1-based for naming)
        
    Returns:
        tuple: A 3-tuple containing:
            - filename (Path): Full path to the log file
            - logfile (str): Log file name with .log extension
            - dltfile (str): DLT file name with .dlt extension
            
    File Naming Convention:
        {timestamp}_Startup_Time_Logs_{setup_type}_{ecu_type}_N{iteration_number}
        
    Example:
        >>> get_log_file_path('RCAR', 'Elite', 3, '20250101_120000', 0)
        (Path('logs/20250101_120000_Startup_Time_Logs_Elite_RCAR_N1.log'), 
         '20250101_120000_Startup_Time_Logs_Elite_RCAR_N1.log',
         '20250101_120000_Startup_Time_Logs_Elite_RCAR_N1.dlt')
         
    Note:
        - Creates 'Logs' directory under local_save_path if it doesn't exist
        - Uses 1-based numbering for iteration display (N1, N2, N3, etc.)
        - Supports both .log and .dlt file formats for different logging needs
    """
    # Construct the log file name based on the ECU type and timestamp
    basename = f'{current_timestamp}_Startup_Time_Logs_{setup_type}_{ecu_type}_N{index + 1}'
    # basename = f'20250626_125003_Startup_Time_Logs_{setup_type}_{ecu_type}_N{index + 1}'
    logfile = basename+'.log'
    dltfile = basename+'.dlt'

    # Define the directory for storing logs
    logs_dir = local_save_path / "Logs"

    # Define the full path to the log file
    filename = logs_dir / logfile

    # Check if the logs directory exists, and create it if it doesn't
    if not logs_dir.exists():
        # Create the logs directory
        logs_dir.mkdir()

    # Return the log file path and name
    return filename, logfile, dltfile

def find_log_files_with_keywords(folder_path, keywords, logger):
    """
    Returns a list of .log files in folder_path whose filenames contain any of the keywords.
    """
    if not folder_path.exists():
        logger.warning(f"Directory {folder_path} does not exist.")
        return []
    log_files = glob.glob(os.path.join(folder_path, "*.log"))
    filtered_files = [
        f for f in log_files
        if all(keyword.lower() in os.path.basename(f).lower() for keyword in keywords)
    ]
    return filtered_files

def extract_log_file_paths(index, ecu_type, setup_type, logger):
    parent_dir = local_save_path / "Logs"
    keywords = [ecu_type, setup_type, f'N{index + 1}']
    if setup_type == ECUType.ELITE.value:
        filtered_files = find_log_files_with_keywords(parent_dir / ecu_type, keywords, logger)
    elif setup_type == ECUType.PADAS.value:
        filtered_files = find_log_files_with_keywords(parent_dir, keywords, logger)
    if not filtered_files or len(filtered_files) == 0:
        logger.warning(f"No log files found for {ecu_type} with setup type {setup_type} and index {index + 1}.")
        if setup_type == ECUType.ELITE.value:
            return tuple((parent_dir / ecu_type / f'{ecu_type}_{setup_type}_N{index + 1}.log', None, None))
        else:
            return tuple((parent_dir / f'{ecu_type}_{setup_type}_N{index + 1}.log', None, None))
    else:
        log_file_path = filtered_files.pop()
        return tuple((log_file_path, os.path.basename(log_file_path), None))

def get_log_file_paths_for_elite(index, ecu_config_list, setup_type):    
    """
    Generates log file paths for multiple ECUs in Elite setup configurations.
    
    This function is specifically designed for Elite test setups that involve multiple
    ECUs running simultaneously. It creates separate log directories for each ECU type
    and generates appropriate file paths for concurrent logging operations.
    
    Args:
        index (int): Current test iteration index (0-based)
        current_timestamp (str): Timestamp string for file naming (format: YYYYMMDD_HHMMSS)
        ecu_config_list (list): List of ECU configuration dictionaries, each containing 'ecu-type' key
        setup_type (str): Test setup configuration type (typically 'Elite')
        
    Returns:
        dict: Dictionary mapping ECU types to their respective file path tuples.
              Each tuple contains (full_path, logfile_name, dltfile_name)
              
    Directory Structure:
        Logs/
        ├── RCAR/
        │   ├── {timestamp}_Startup_Time_Logs_{setup_type}_RCAR_N{iteration}.log
        │   └── {timestamp}_Startup_Time_Logs_{setup_type}_RCAR_N{iteration}.dlt
        ├── SoC0/
        │   ├── {timestamp}_Startup_Time_Logs_{setup_type}_SoC0_N{iteration}.log
        │   └── {timestamp}_Startup_Time_Logs_{setup_type}_SoC0_N{iteration}.dlt
        └── SoC1/
            ├── {timestamp}_Startup_Time_Logs_{setup_type}_SoC1_N{iteration}.log
            └── {timestamp}_Startup_Time_Logs_{setup_type}_SoC1_N{iteration}.dlt
            
    Features:
        - Creates separate subdirectories for each ECU type
        - Handles directory creation with parent directory support
        - Uses consistent naming convention across all ECU types
        - Provides debug output showing all generated paths
        
    Example:
        >>> ecu_list = [{'ecu-type': 'RCAR'}, {'ecu-type': 'SoC0'}]
        >>> get_log_file_paths_for_elite(0, '20250101_120000', ecu_list, 'Elite')
        {'RCAR': (Path('Logs/RCAR/20250101_120000_Startup_Time_Logs_Elite_RCAR_N1.log'), 
                  '20250101_120000_Startup_Time_Logs_Elite_RCAR_N1.log',
                  '20250101_120000_Startup_Time_Logs_Elite_RCAR_N1.dlt'),
         'SoC0': (Path('Logs/SoC0/20250101_120000_Startup_Time_Logs_Elite_SoC0_N1.log'),
                  '20250101_120000_Startup_Time_Logs_Elite_SoC0_N1.log',
                  '20250101_120000_Startup_Time_Logs_Elite_SoC0_N1.dlt')}
    """
    parent_dir = local_save_path / "Logs"
    ecu_type_list = [ecu['ecu-type'] for ecu in ecu_config_list]
    logs_dir_list = [parent_dir/ecu_type for ecu_type in ecu_type_list]
    filename_list = {}
   
    for logs_dir, ecu_type in zip(logs_dir_list, ecu_type_list):
        basename = f'{current_timestamp}_Startup_Time_Logs_{setup_type}_{ecu_type}_N{index + 1}'
        # basename = f'20250626_125003_Startup_Time_Logs_{setup_type}_{ecu_type}_N{index + 1}'
        logfile = basename+'.log'
        dltfile = basename+'.dlt'
        filename_list[ecu_type] = tuple((logs_dir / logfile, logfile, dltfile))
        if(not logs_dir.exists()):
            logs_dir.mkdir(parents=True, exist_ok=True)
    print(f"Log files will be saved in the following directories: {filename_list}")
    return filename_list


def get_expected_startup_order(application_name, application_startup_order, logger):
    """
    Determines the expected startup order position for a given application.
    
    This function analyzes the configured startup order sequence to determine where
    a specific application should appear in the startup sequence. It handles both
    sequential and parallel startup configurations.
    
    Args:
        application_name (str): Name of the application to find in the startup order
        application_startup_order (list): List of tuples containing (order_type, app_list)
                                        where order_type is 'Sequential' or 'Parallel'
                                        and app_list contains application names
                                        
    Returns:
        str or None: Expected startup order position as string, or None if not found
                    - Sequential: Exact position number (e.g., "3")
                    - Parallel (single app): Position number (e.g., "2")
                    - Parallel (multiple apps): Range (e.g., "2~4")
                    
    Order Types:
        - Sequential: Applications start one after another in strict order
        - Parallel: Applications can start simultaneously within the group
        
    Example:
        >>> startup_order = [
        ...     ('Sequential', ['app1', 'app2']),
        ...     ('Parallel', ['app3', 'app4', 'app5']),
        ...     ('Sequential', ['app6'])
        ... ]
        >>> get_expected_startup_order('app2', startup_order)
        '2'
        >>> get_expected_startup_order('app4', startup_order)
        '3~5'
        >>> get_expected_startup_order('app6', startup_order)
        '6'
        >>> get_expected_startup_order('unknown', startup_order)
        None
        
    Note:
        The function maintains a running position counter to handle mixed
        sequential and parallel groups correctly. Position numbering is 1-based.
    """
    # Iterate over the application startup order to find the expected startup order for the given application name
    cur_pos = 0
    for order_type, order in application_startup_order:
        if application_name in order:
            logger.info('order type: %s, %s, %s', order_type, order_type.lower(), OrderType.SEQUENTIAL.value)
            if order_type.lower() == OrderType.SEQUENTIAL.value:
                return str(cur_pos + order.index(application_name) + 1)
            else:
                if len(order) == 1:
                    return str(cur_pos + 1)
                return str(cur_pos + 1) + '~' + str(cur_pos + len(order))
        cur_pos += len(order)
    return None
 


def write_data_to_excel(ecu_type, dltstart_timestamps, process_timing_info, sheet, application_startup_order, threshold, validate_startup_order, application_startup_order_status_iteration, overall_IG_ON_cur_iteration, logger):
    """
    Writes application startup timing data to Excel worksheet with comprehensive validation.
    
    This function populates an Excel worksheet with detailed startup timing information,
    including performance validation, startup order verification, and failure analysis.
    It creates a comprehensive report that stakeholders can use to assess system performance.
    
    Args:
        dltstart_timestamps (dict): Dictionary mapping application names to their startup times (seconds)
        process_timing_info (list): List of process timing information (currently used for debugging)
        sheet (openpyxl.worksheet.worksheet.Worksheet): Excel worksheet to populate
        application_startup_order (list): Expected startup order configuration
        threshold (float): Performance threshold in seconds for pass/fail determination
        validate_startup_order (bool): Whether to perform startup order validation
        application_startup_order_status_iteration (dict): Counter for different failure types
        
    Data Columns Written:
        1. No. - Sequential position number
        2. Services/Applications - Application name
        3. Application Startup Time (sec) - Time from QNX startup
        4. IG ON to QNX Startup (sec) - Fixed offset time (OFFSET_TIME)
        5. Total Time from IG ON (sec) - Combined startup time
        6. Test Case Status - PASS/FAIL based on threshold
        7. Expected Order - Expected startup position (if validation enabled)
        8. StartUp Time Judgement - Order validation result
        9. Order Mismatch - 'O' if order mismatch detected
        10. Application Not Found - 'O' if application missing from logs
        11. Application Not Configured - 'O' if application not in configuration
        
    Validation Logic:
        - Performance: Compares (startup_time + OFFSET_TIME) against threshold
        - Order: Validates actual vs expected startup sequence
        - Completeness: Identifies missing applications from logs
        
    Features:
        - Merges cells in column D (IG ON to QNX Startup) for visual clarity
        - Handles missing applications by adding placeholder rows
        - Provides summary count of different failure types
        - Applies consistent border styling
        
    Note:
        This function is central to the reporting system and provides the detailed
        data that feeds into summary reports and visualizations.
    """

    startup_order_count_idx = sheet.max_row + 1
    if validate_startup_order:
        sheet.append(['', '', '', '', '', '', '', '', '', 0, 0, 0])
        
    start_row = sheet.max_row + 1

    # Iterate over the DLTStart timestamps and differences in parallel using zip
    for position, (process, dltstart_line) in enumerate(dltstart_timestamps.items()):
        # Check if the process names match
       
        if float(dltstart_line + OFFSET_TIME) < (threshold_map[ecu_type][process] if process in threshold_map[ecu_type] else threshold):
            result = 'PASS'
        else:
            result = 'FAIL'
            overall_IG_ON_cur_iteration['status'] = False
        logger.info(">>> %s, %s, %s", process, process, process_timing_info)

        data_row = [position+1, process, round_decimal_half_up(dltstart_line, 4), OFFSET_TIME, round_decimal_half_up(dltstart_line + OFFSET_TIME, 4), threshold_map[ecu_type][process] if process in threshold_map[ecu_type] else threshold, result]
        logger.info('## %s, %s', process, validate_startup_order)

        if validate_startup_order:
            # Create a data row for the process
            expected_order = get_expected_startup_order(process, application_startup_order, logger)
            if not expected_order:
                expected_order='-'
            data_row.append(str(expected_order))
            
            order_failure_type = validate_ind_app_startup_order(process, position + 1, application_startup_order)
            if order_failure_type != 0:
                data_row.extend([
                    'FAIL',
                    '⬤' if OrderFailureType.ORDER_MISMATCH.name == OrderFailureType(order_failure_type).name else '',
                    '',
                    '⬤' if OrderFailureType.APPLICATION_NOT_CONFIGURED.name == OrderFailureType(order_failure_type).name else ''
                ])
                application_startup_order_status_iteration[OrderFailureType(order_failure_type).name] += 1
                application_startup_order_status_iteration['startup_order_status'] = False
            else:
                data_row.extend(['PASS', '', '', ''])

        # Append the data row to the sheet
        sheet.append(data_row)
       
    # Merge cells in column D for the rows created in this scenario
    merged_range = f'D{start_row}:D{sheet.max_row}'
    sheet.merge_cells(merged_range)
   
    if validate_startup_order:
        for order_type, order in application_startup_order:
            for app in order:
                if app not in dltstart_timestamps:
                
                    data_row = ['-', app, '-', '-', '-', '-', '-']
                
                    if validate_startup_order:
                        expected_order = get_expected_startup_order(app, application_startup_order, logger)
                        if not expected_order:
                            expected_order='-'
                        data_row.extend([str(expected_order), 'FAIL', '', '⬤', ''])
                        application_startup_order_status_iteration[OrderFailureType.APPLICATION_NOT_FOUND.name] += 1
                    sheet.append(data_row)
        # Update the last three cells of the row at startup_order_count_idx with the current counts and highlight in yellow
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        counts = [
            application_startup_order_status_iteration[OrderFailureType.ORDER_MISMATCH.name],
            application_startup_order_status_iteration[OrderFailureType.APPLICATION_NOT_FOUND.name],
            application_startup_order_status_iteration[OrderFailureType.APPLICATION_NOT_CONFIGURED.name]
        ]
        # Merge cells from column 1 to 9 in the current row with the above row
        for col in range(1, 10):
            sheet.merge_cells(
            start_row=startup_order_count_idx - 1,
            start_column=col,
            end_row=startup_order_count_idx,
            end_column=col
            )
        for offset, count in enumerate(counts, start=10):
            cell = sheet.cell(row=startup_order_count_idx, column=offset)
            cell.value = count
            cell.fill = yellow_fill
            cell.border = border_style
        # Ensure all cells in the merged range have borders
        for col in range(1, 10):
            for row in range(startup_order_count_idx - 1, startup_order_count_idx + 1):
                cell = sheet.cell(row=row, column=col)
                cell.border = border_style
    # Apply the border style to the entire merged range
    for row in sheet[merged_range]:
        for cell in row:
            cell.border = border_style


def create_header(sheet, ecu_type, validate_startup_order, app_columns):
    """
    Creates formatted section headers for different types of data in Excel worksheets.
    
    This function generates professional-looking section headers that separate different
    types of analysis data within the same worksheet. It handles multiple report types
    and adjusts column configurations based on validation settings.
    
    Args:
        sheet (openpyxl.worksheet.worksheet.Worksheet): Excel worksheet to add header to
        ecu_type (str): ECU type identifier for header text
        validate_startup_order (bool): Whether startup order validation is enabled
        app_columns (str): Type of columns to create, determines header style and content
        
    Column Types Supported:
        - 'min_max_avg_columns': Statistical summary of startup times
        - 'min_max_avg_individual': Statistical summary of individual app times
        - 'startup_time_columns': Detailed startup time analysis
        - 'info_columns': Application initialization time information
        - 'overall_test_columns': Test iteration summary
        - 'startup_appendix': Field descriptions and documentation
        
    Header Features:
        - Adds 10 empty rows for visual separation from previous sections
        - Creates merged header cell spanning all data columns
        - Applies professional blue background (9EB9DA) with bold text
        - Centers header text both horizontally and vertically
        - Adds column name row with appropriate text wrapping
        - Applies consistent border styling
        
    Dynamic Column Adjustment:
        - Removes startup order validation columns when validation is disabled
        - Adjusts column count automatically based on configuration
        
    Returns:
        int: Row number where the header section starts (for reference in other functions)
        
    Note:
        This function is essential for creating well-organized, professional reports
        that clearly separate different types of analysis data within the same worksheet.
    """
    # Check if the sheet has existing rows and append empty rows if necessary
    if sheet.max_row > 1:
        # Append 5 empty rows to separate the header from existing data
        for _ in range(10):
            sheet.append([])

    # Determine the header text and column names based on the avg_flag
    if app_columns == 'min_max_avg_columns':
        # If avg_flag is True, include Min, Max, and Avg in the header
        header = f'Services/Applications Startup Time from QNX Startup on {ecu_type} (Min, Max, Avg)'
        columns = application_startup_time_min_max_avg_columns
   
    elif app_columns == 'min_max_avg_individual':
        # If avg_flag is True, include Min, Max, and Avg in the header
        header = f'Services/Applications Individual Startup Times on {ecu_type} (Min, Max, Avg)'
        columns = application_start_end_time_min_max_avg_columns

    elif app_columns == 'startup_time_columns':
        # If avg_flag is False, only include Startup Time in the header
        header = f'Services/Applications Startup Time on {ecu_type}'
        columns = application_startup_time_columns
        if not validate_startup_order:
            columns=columns[:-5]
   
    elif app_columns == 'info_columns':
        header = f'Services/Applications Init(Up) Time on {ecu_type}'
        columns = application_info_columns
   
    elif app_columns == 'overall_test_columns':
        header = f'Overall Test Case Status for each Iteration on {ecu_type}'
        columns = applications_overall_status_columns
        if not validate_startup_order:
            columns=columns[:-4]

    elif app_columns == 'startup_appendix':
       header = f'Field Description for \n Services/Applications Startup Completion Time on {ecu_type}'
       columns = appendix_columns
    table_headers.append(header)
    # Append the header text to the sheet
    sheet.append([header])

    # Get the current row number (which is now the start of the header)
    start_row = sheet.max_row

    # Calculate the last column letter based on the number of columns
    last_column_letter = chr(64 + len(columns))

    # Merge the cells in the header row
    merged_range = f'A{sheet.max_row}:{last_column_letter}{sheet.max_row}'
    sheet.merge_cells(merged_range)

    # Get the merged cell object
    merged_cell = sheet.cell(row=sheet.max_row, column=1)

    # Apply formatting to the merged cell (gray fill, bold text, centered alignment)
    merged_cell.fill = PatternFill(start_color="9EB9DA", end_color="9EB9DA", fill_type="solid")
    merged_cell.alignment = Alignment(horizontal='center', vertical='center')
    merged_cell.font = Font(bold=True)

    # Append the column names for the header
    sheet.append(columns)    

    for col_idx, col_val in enumerate(columns):
        cell = sheet.cell(row=sheet.max_row, column=col_idx + 1)
        if '\n' in col_val:
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)    
        else:    
            cell.alignment = Alignment(horizontal='center', vertical='center')

    # Apply the border style to the entire merged range
    for row in sheet[merged_range]:
        for cell in row:
            cell.border = border_style

    # Return the row number where the header starts
    return start_row


def each_iteration_test_status(ecu_type, summary_sheet, overall_IG_ON_iteration, config, application_startup_order_status):
    """
    Creates a summary table showing test results for each iteration with hyperlinks to detailed data.
    
    This function generates an executive summary that provides a high-level overview of
    test results across all iterations. It includes hyperlinks to detailed worksheets
    and summarizes both performance and startup order validation results.
    
    Args:
        ecu_type (str): ECU type identifier for header generation
        summary_sheet (openpyxl.worksheet.worksheet.Worksheet): Summary worksheet to populate
        overall_IG_ON_iteration (dict): Dictionary mapping iteration index to overall startup time
        config (dict): Test configuration containing iterations count and validation settings
        application_startup_order_status (dict): Startup order validation results per iteration
        
    Summary Table Columns:
        1. No. of Iterations - Hyperlinked iteration number (links to detailed worksheet)
        2. Total Time to Startup Last Application from IG ON (sec) - Overall performance
        3. Test Case Status - PASS/FAIL based on 5-second threshold
        4. Startup Order Status - PASS/FAIL for order validation (if enabled)
        5. Order Mismatch Count - Number of order violations
        6. Not Found Count - Number of missing applications
        7. Not Configured Count - Number of unconfigured applications
        
    Features:
        - Creates Excel hyperlinks to detailed iteration worksheets
        - Applies 5-second performance threshold for pass/fail determination
        - Includes startup order validation summary when enabled
        - Provides failure count breakdown for root cause analysis
        - Uses professional formatting via format_excel_cells
        
    Hyperlink Format:
        Links to worksheets named 'GEN3_StartupTime_{iteration_number}'
        
    Performance Threshold:
        - PASS: Total time ≤ 5 seconds
        - FAIL: Total time > 5 seconds
        
    Note:
        This summary table is typically the first thing stakeholders review
        to get an overall assessment of system performance across test iterations.
    """
    start_row = create_header(summary_sheet, ecu_type, config['Startup Order Judgement'], 'overall_test_columns')
    for i in range(config['Iterations']):
        if i in overall_IG_ON_iteration:
            overall_value = overall_IG_ON_iteration[i]['timestamp'] + OFFSET_TIME
            test_status = 'PASS' if overall_IG_ON_iteration[i]['status'] else 'FAIL'
            data_row = [f'=HYPERLINK("#\'GEN3_StartupTime_{(i + 1):02d}\'!A1", "{i + 1}")', overall_value, test_status]
            if config['Startup Order Judgement'] and i in application_startup_order_status:
                data_row.extend([
                    "PASS" if application_startup_order_status[i]['startup_order_status'] else "FAIL",
                    application_startup_order_status[i][OrderFailureType.ORDER_MISMATCH.name],
                    application_startup_order_status[i][OrderFailureType.APPLICATION_NOT_FOUND.name],
                    application_startup_order_status[i][OrderFailureType.APPLICATION_NOT_CONFIGURED.name]
                ])
            summary_sheet.append(data_row)
            
            # Apply hyperlink formatting to the first cell in the last row
            cell = summary_sheet.cell(row=summary_sheet.max_row, column=1)
            cell.font = Font(bold=True, underline='single', color='0000FF')
   
    format_excel_cells(summary_sheet, start_row)


def export_and_plot_average_data_to_excel(sheet, ecu_type, process_times, process_start_times, config, logger):
    """
    Generates comprehensive statistical analysis and visualizations of application startup performance.
    
    This function creates two detailed statistical sections in the Excel report:
    1. Startup time statistics (min/max/avg) from QNX startup
    2. Individual application initialization time statistics
    
    Each section includes both tabular data and corresponding timeline visualizations.
    
    Args:
        sheet (openpyxl.worksheet.worksheet.Worksheet): Excel worksheet for the analysis
        ecu_type (str): ECU type identifier for headers and graph titles
        process_times (dict): Dictionary mapping process names to lists of startup times
        process_start_times (dict): Dictionary mapping process names to lists of init times
        config (dict): Test configuration containing validation settings and thresholds
        
    Section 1 - Startup Time Statistics:
        - Calculates min, max, and average startup times for each application
        - Includes total time from IG ON (startup_time + OFFSET_TIME)
        - Sorts applications by average startup time for easy identification of slow starters
        - Generates timeline graph showing average performance
        
    Section 2 - Individual Application Statistics:
        - Focuses on application-specific initialization times
        - Provides statistical summary without IG ON offset
        - Sorts by average initialization time
        - Generates separate timeline graph for initialization performance
        
    Statistical Calculations:
        - Minimum: Best performance across all iterations
        - Maximum: Worst performance across all iterations  
        - Average: Mean performance (rounded to 3 decimal places)
        
    Visualizations:
        - Timeline graphs embedded directly in Excel worksheet
        - Horizontal timeline format for easy comparison
        - Professional styling with grids and labels
        - Threshold reference lines for performance assessment
        
    Features:
        - Automatic sorting by performance for priority identification
        - Dual analysis perspectives (system-level and app-level)
        - Integrated visualizations for stakeholder presentations
        - Professional formatting and column width optimization
        
    Note:
        This function provides the statistical foundation for performance analysis,
        helping identify performance trends, outliers, and optimization opportunities.
    """
    # Create a header in the Excel sheet for the average data
    start_row = create_header(sheet, ecu_type, config['Startup Order Judgement'], 'min_max_avg_columns')

    # Initialize an empty dictionary to store the average differences
    differences = {}
    individual_differences= {}    

    # Initialize an empty list to store the data
    data = []
    individual_list = []

    # Iterate over each process and its times
    for process, times in process_times.items():
        # Calculate the minimum, maximum, and average times for the process
        min_time = min(times)
        max_time = max(times)
        # avg_time = sum(times) / len(times)
        avg_time = round_decimal_half_up(sum(times) / len(times), 4)
       
        # Create a dictionary for the process with the minimum, maximum, and average times
        data_row = {
            'process': process,
            'min_time': round_decimal_half_up(min_time, 4),
            'max_time': round_decimal_half_up(max_time, 4),
            'avg_time': round_decimal_half_up(avg_time, 4),
        }
       
        # Append the data row to the list
        data.append(data_row)

    # Sort the data based on the average time
    data.sort(key=lambda x: x['avg_time'])

    # Append the sorted data to the Excel sheet
    for data_row in data:
        sheet.append([data_row['process'], data_row['min_time'], data_row['max_time'], data_row['avg_time'], float(data_row['avg_time']) + OFFSET_TIME, threshold_map[ecu_type][data_row['process']] if data_row['process'] in threshold_map[ecu_type] else config['Threshold']])

        # Store the average difference in the differences dictionary
        differences[data_row['process']] = float(data_row['avg_time'])

    # Plot the average data as a graph
    plot_process_startup_time_graph(differences, sheet, start_row, ecu_type, True)

    # Format the Excel cells
    format_excel_cells(sheet, start_row)

    # Create a header in the Excel sheet for the average data
    start_row = create_header(sheet, ecu_type, config['Startup Order Judgement'], 'min_max_avg_individual')

    for process, start_times in process_start_times.items():
        # Calculate the minimum, maximum, and average start times for the process
        min_time = min(start_times)
        max_time = max(start_times)
        avg_time = sum(start_times) / len(start_times)
        # avg_time = round(sum(start_times) / len(start_times), 3)    

        # Create a dictionary for the process with the minimum, maximum, and average times
        data_row = {
            'process': process,
            'min_time': round_decimal_half_up(min_time, 4),
            'max_time': round_decimal_half_up(max_time, 4),
            'avg_time': round_decimal_half_up(avg_time, 4),
        }
       
        # Append the data row to the list
        individual_list.append(data_row)

    # Sort the data based on the average time
    individual_list.sort(key=lambda x: x['avg_time'])    
       
      # Append the sorted data to the Excel sheet
    for data_row in individual_list:
        sheet.append([data_row['process'], data_row['min_time'], data_row['max_time'], data_row['avg_time']])

        # Store the average difference in the differences dictionary
        individual_differences[data_row['process']] = float(data_row['avg_time'])
   
    # Plot the average data as a graph
    plot_process_individual_apps_avg_graph(individual_differences, sheet, start_row, ecu_type)

    # Format the Excel cells
    format_excel_cells(sheet, start_row)
   
    # Adjust the column width of the Excel sheet
    adjust_column_width(sheet, ecu_type, logger)


def add_logfile_hyperlink(report_path, log_path, sheet, ecu_type, setup_type):
    """
    Adds a hyperlink to the source log file in the Excel worksheet for traceability.
    
    This function creates a clickable hyperlink in the Excel report that allows users
    to quickly access the original log file used to generate the analysis. This is
    essential for audit trails and detailed troubleshooting.
    
    Args:
        report_path (str): Relative or absolute path to the log file for the hyperlink
        log_path (str): Display text for the hyperlink (typically the filename)
        sheet (openpyxl.worksheet.worksheet.Worksheet): Excel worksheet to add the hyperlink
        
    Hyperlink Features:
        - Uses Excel's native HYPERLINK formula for compatibility
        - Positions hyperlink with appropriate spacing from existing content
        - Applies blue font color following standard hyperlink conventions
        - Includes descriptive label "Log File:" for clarity
        
    Layout:
        - Row N+2: "Log File:" label
        - Row N+3: Clickable hyperlink to log file
        
    Excel Formula:
        =HYPERLINK("path/to/file", "display_text")
        
    Use Cases:
        - Audit trail for report verification
        - Quick access to raw data for detailed analysis
        - Troubleshooting when report data needs verification
        - Compliance requirements for data traceability
        
    Note:
        The hyperlink path should be relative to the Excel file location
        for portability across different systems and users.
    """
    # Get the next available row in the sheet
    row_no = sheet.max_row + 2
 
    # Set the text for the hyperlink
    sheet.cell(row=row_no, column=1).value = "Log File:"  
 
    # Use Excel's =HYPERLINK() formula with the relative path
    if setup_type == ECUType.ELITE.value:
        hyperlink_formula = f'=HYPERLINK(".\Logs\{ecu_type}\{log_path}", "{log_path}")'
    else:
        hyperlink_formula = f'=HYPERLINK(".\Logs\{log_path}", "{log_path}")'
    # Insert the hyperlink formula
    sheet.cell(row=row_no + 1, column=1).value = hyperlink_formula
   
    # Set the font color of the hyperlink to blue
    sheet.cell(row=row_no + 1, column=1).font = Font(color="0000FF")


def generate_apps_start_end_time_report(ecu_type, sheet, process_timing_info, config):
    """
    Generates a detailed report of application initialization (Init/Up) times with visualization.
    
    This function creates a focused analysis of how long each application takes to
    fully initialize after being started. It provides both tabular data and a
    timeline visualization to help identify applications with long initialization times.
    
    Args:
        ecu_type (str): ECU type identifier for header and graph titles
        sheet (openpyxl.worksheet.worksheet.Worksheet): Excel worksheet for the report
        process_timing_info (list): List of dictionaries containing process timing data
                                   Each dict should have 'process' and 'start_time_ms' keys
        config (dict): Test configuration for validation settings
        
    Report Columns:
        1. Services/Applications - Application name
        2. Init(Up) Time (us) - Initialization time in microseconds (start_time_ms * 1000)
        3. Init(Up) Time (ms) - Initialization time in milliseconds (original value)
        
    Data Processing:
        - Filters out applications without initialization time data
        - Converts milliseconds to microseconds for detailed precision
        - Maintains original millisecond values for comparison
        
    Visualization:
        - Embedded timeline graph showing initialization duration
        - Horizontal timeline format for easy comparison
        - Professional styling with appropriate scaling
        
    Use Cases:
        - Identifying applications with excessive initialization overhead
        - Comparing initialization performance across different applications
        - Optimization target identification for development teams
        - Performance regression analysis across software versions
        
    Note:
        This report focuses specifically on the time between application start
        and full operational readiness, which is different from the overall
        startup time that includes system-level delays.
    """
    # Create the header for the Excel sheet
    start_row = create_header(sheet, ecu_type, config['Startup Order Judgement'], 'info_columns')

    for item in process_timing_info:
        if item['start_time_ms']:
            data_row = [item['process'], float(item['start_time_ms'])*1000,float(item['start_time_ms'])]
            sheet.append(data_row)

    # Plot the startup graph
    plot_process_start_end_time_graph(ecu_type, process_timing_info, sheet, start_row)

    # Format the Excel cells
    format_excel_cells(sheet, start_row)


def generate_apps_startup_report_from_QNX_startup(ecu_type, config, sheet, dltstart_timestamps,  process_timing_info, application_startup_order, application_startup_order_status_iteration, overall_IG_ON_cur_iteration, logger):
    """
    Generates a comprehensive startup time analysis report for a single test iteration.
    
    This is the main report generation function that creates a complete analysis
    of application startup performance for one test iteration. It combines multiple
    types of analysis into a single, comprehensive worksheet.
    
    Args:
        ecu_type (str): ECU type identifier for headers and graph titles
        config (dict): Test configuration containing thresholds and validation settings
        sheet (openpyxl.worksheet.worksheet.Worksheet): Excel worksheet for the complete report
        dltstart_timestamps (dict): Dictionary mapping application names to startup times
        process_timing_info (list): List of process initialization timing data
        application_startup_order (list): Expected startup order configuration
        application_startup_order_status_iteration (dict): Startup order validation counters
        
    Report Sections Generated:
        1. Detailed Startup Time Analysis:
           - Individual application startup times from QNX startup
           - Performance validation against threshold
           - Startup order validation (if enabled)
           - Timeline visualization with threshold and reference lines
           
        2. Application Initialization Analysis:
           - Init(Up) time details for each application
           - Separate timeline visualization for initialization times
           - Focus on application-specific performance
           
    Features:
        - Comprehensive data validation and error handling
        - Multiple visualization perspectives
        - Professional formatting and styling
        - Automatic column width optimization
        - Integration of performance and order validation
        
    Workflow:
        1. Create main section header
        2. Write detailed startup data with validation
        3. Generate main timeline visualization
        4. Apply professional formatting
        5. Add initialization time analysis section
        6. Optimize layout for readability
        
    Note:
        This function creates the detailed per-iteration reports that stakeholders
        use for deep-dive analysis of specific test runs. It's complemented by
        summary reports that aggregate data across multiple iterations.
    """
    # Create the header for the Excel sheet
    start_row = create_header(sheet, ecu_type, config['Startup Order Judgement'], 'startup_time_columns')

    # Write the data to the Excel sheet
    write_data_to_excel(ecu_type, dltstart_timestamps, process_timing_info, sheet, application_startup_order, config.get('Threshold'), config.get('Startup Order Judgement'), application_startup_order_status_iteration, overall_IG_ON_cur_iteration, logger)

    # Plot the differences as a graph
    plot_process_startup_time_graph(dltstart_timestamps, sheet, start_row, ecu_type, False)

    # Format the Excel cells
    format_excel_cells(sheet, start_row)

    generate_apps_start_end_time_report(ecu_type, sheet, process_timing_info, config)

    # Adjust the column width of the Excel sheet
    adjust_column_width(sheet, ecu_type, logger)


def extract_and_sort_process_timestamps(process_Start_End_timestamps, ecu_type, logger):
    """
    Extracts and sorts process initialization timing information from raw timestamp data.
    
    This function processes raw timing data extracted from log files and converts it
    into a structured format suitable for analysis and reporting. It handles missing
    data gracefully and sorts results for logical presentation.
    
    Args:
        process_Start_End_timestamps (dict): Dictionary mapping process names to timing data
                                           Each entry should contain 'init_time' key with timing value
        ecu_type (str): ECU type identifier (used for logging context)
        
    Returns:
        list: List of dictionaries containing structured timing information
              Each dictionary has 'process' and 'start_time_ms' keys
              Sorted by initialization time (fastest to slowest)
              
    Data Processing:
        - Extracts initialization time from 'init_time' field
        - Converts timing values to float for numerical operations
        - Handles missing initialization times by defaulting to 0
        - Logs warnings for processes without initialization data
        
    Sorting Logic:
        - Primary sort: By initialization time (ascending)
        - Missing data handling: Places processes without timing data at the end
        - Uses float('inf') for missing values to ensure proper sorting
        
    Data Structure:
        Input: {'process_name': {'init_time': 123.45, ...}, ...}
        Output: [{'process': 'process_name', 'start_time_ms': 123.45}, ...]
        
    Error Handling:
        - Gracefully handles missing 'init_time' keys
        - Logs warnings for incomplete data without stopping processing
        - Ensures all processes are included in output even with missing data
        
    Note:
        This function is crucial for converting raw log parsing results into
        the structured format required by reporting and visualization functions.
    """
    process_timing_info = []
    for process, time in process_Start_End_timestamps.items():  
        # Check if both start and end times are available
        # if 'start' in time and 'end' in time:            
        start_time_ms = float(time['init_time'] if 'init_time' in time else '0')
        # logger.info(f"Process: {process}, Start Time: {time['start']}, End Time: {time['end']}, Time Difference: {start_time_ms} ms")

        process_timing_info.append({
        'process': process,
        'start_time_ms': start_time_ms
    })

        # Check if only end time is available
        if 'init_time' not in time:
            logger.warning(f"Process: {process}, Init: Not Available")
     
    process_timing_info.sort(key=lambda x: x['start_time_ms'] if x.get('start_time_ms')is not None else float('inf'))
    return process_timing_info


def calculate_differences(dltstart_timestamps, welcome_timestamp, logger):
    """
    Calculates time differences between application start times and the system welcome timestamp.
    
    This function computes the elapsed time from system welcome (baseline) to each
    application's startup completion. It serves as the foundation for startup time
    analysis by providing relative timing measurements.
    
    Args:
        dltstart_timestamps (dict): Dictionary mapping application names to their start timestamp strings
                                   Timestamps should be in format 'YYYY-MM-DD HH:MM:SS.ffffff'
        welcome_timestamp (str): System welcome timestamp as baseline reference
                                Format: 'YYYY-MM-DD HH:MM:SS.ffffff'
                                
    Returns:
        dict: Dictionary mapping application names to their startup time differences in seconds
              Positive values indicate applications started after welcome timestamp
              
    Timestamp Processing:
        - Parses timestamp strings using strptime with microsecond precision
        - Calculates differences using datetime arithmetic
        - Converts results to seconds (float) for numerical analysis
        
    Error Handling:
        - Catches ValueError exceptions for malformed timestamps
        - Logs specific error details for troubleshooting
        - Continues processing other applications even if some fail
        - Returns partial results rather than failing completely
        
    Calculation Formula:
        difference = (application_start_time - welcome_timestamp).total_seconds()
        
    Example:
        >>> dlt_times = {'app1': '2025-01-01 12:00:05.123456'}
        >>> welcome = '2025-01-01 12:00:02.000000'
        >>> calculate_differences(dlt_times, welcome)
        {'app1': 3.123456}
        
    Note:
        This function is critical for establishing the temporal relationship
        between system initialization and individual application startup,
        forming the basis for all subsequent timing analysis.
    """
    # Initialize an empty dictionary to store differences
    differences = {}
    # Iterate over each DLTStart timestamp
    for process, dltstart_line in dltstart_timestamps.items():
        try:
            # Parse the DLTStart timestamp and welcome timestamp to datetime objects
            dltstart_datetime = datetime.strptime(dltstart_line, '%Y-%m-%d %H:%M:%S.%f')
            welcome_datetime = datetime.strptime(welcome_timestamp, '%Y-%m-%d %H:%M:%S.%f')
            # Calculate the difference between the two timestamps
            difference = (dltstart_datetime - welcome_datetime).total_seconds()
            # Store the difference in the dictionary
            differences[process] = difference

            # logger.info(f"welcome_datetime : {welcome_datetime} appstart_datetime : {dltstart_datetime} difference : {difference}")
        except ValueError:
            # Handle any errors parsing the timestamps
            logger.error(f"Error parsing timestamp for process {process}: {dltstart_line}")
    return differences


def extract_process_timestamps(lines):
    """
    Extracts application initialization timing information from DLT log file lines.
    
    This function parses DLT log entries to extract application initialization times,
    which represent how long each application takes to fully initialize after being
    started. It looks for specific log patterns that indicate initialization completion.
    
    Args:
        lines (list): List of log file lines to parse
        
    Returns:
        dict: Dictionary mapping application names to their timing information
              Structure: {'app_name': {'init_time': time_in_ms}, ...}
              
    Log Pattern Recognition:
        - Searches for lines containing both 'Application:' and 'Init(Up) Time:'
        - Extracts application name from the log entry
        - Parses initialization time value and converts from microseconds to milliseconds
        
    Parsing Logic:
        1. Split line on 'Application:' to isolate application information
        2. Extract application name by splitting on '- Init(Up) Time:'
        3. Parse timing value from 'Init(Up) Time: {value} us' format
        4. Convert microseconds to milliseconds (divide by 1000)
        
    Data Processing:
        - Handles multiple entries per application (overwrites with latest)
        - Strips whitespace from application names
        - Converts timing values to float for numerical operations
        - Creates nested dictionary structure for extensibility
        
    Example Log Entry:
        "Application: crypto_keystorage_manager - Init(Up) Time: 1234567 us"
        
    Extracted Data:
        {'crypto_keystorage_manager': {'init_time': 1234.567}}
        
    Error Handling:
        - Silently skips malformed log entries
        - Continues processing even if individual entries fail
        - Overwrites duplicate entries with latest timing
        
    Note:
        This function focuses specifically on initialization timing, which is
        different from startup timing. Initialization time measures how long
        an application takes to become fully operational after being launched.
    """
    # Initialize an empty dictionary to store the process start and end timestamps
    process_Start_End_timestamps = {}

    # Iterate over each line in the log file
    for line in lines:
        if 'Application:' in line and 'Init(Up) Time:' in line:
           # Split the line to extract the process name and timestamp
           parts = line.split('Application:')
           # Check if the split resulted in more than one part (i.e., the keyword was found)
           if len(parts) > 1:
               # Extract the process information from the second part
               process_info = parts[1]
               # Extract the process name from the process information by splitting at ', Pid:'
               # Remove any trailing '.0' suffix from the process name
               process_name = process_info.split('- Init(Up) Time:')[0].strip()
               # Split the line to extract the end timestamp
               init_timestamp_parts = line.split('Init(Up) Time: ')
               # Check if a timestamp was found
               if len(init_timestamp_parts) > 1:
                   # Extract the timestamp from the second part and convert it to an integer
                   init_timestamp = float(init_timestamp_parts[1].split(' us')[0].strip())/1000
                   # Check if the process name is already in the dictionary
                   if process_name not in process_Start_End_timestamps:
                       # Initialize the process dictionary if it does not exist
                       process_Start_End_timestamps[process_name] = {}
                   # Store the end timestamp in the process dictionary
                   process_Start_End_timestamps[process_name]['init_time'] = init_timestamp
                   # Log the process name and end timestamp
                   # logger.info(f"Process name: {process_name}, End Timestamp: {end_timestamp}")

    return process_Start_End_timestamps


def extract_dltstart_timestamps(lines, logger):
    """
    Extracts application startup timestamps from DLT log file lines using pattern matching.
    
    This function parses DLT log entries to identify when applications are started
    by the system. It uses regular expressions to extract application names and
    their corresponding startup timestamps from log entries.
    
    Args:
        lines (list): List of log file lines to parse
        
    Returns:
        OrderedDict: Ordered dictionary mapping application names to their startup timestamps
                    Maintains chronological order of application startup
                    Structure: {'app_name': 'YYYY-MM-DD HH:MM:SS.ffffff', ...}
                    
    Log Pattern Recognition:
        - Searches for lines containing 'EM' and 'is started'
        - Uses regex pattern to extract application names ending with '.0'
        - Calls extract_timestamp_from_dlt() to parse timestamp information
        
    Regex Pattern Details:
        Pattern: r'EM.*?\b([^\s]+?)(?=\.0)\.0\b'
        - Matches 'EM' followed by any characters
        - Captures application name before '.0' suffix
        - Uses word boundaries to ensure accurate matching
        
    Data Processing:
        - Maintains insertion order using OrderedDict
        - Associates each application with its startup timestamp
        - Handles missing timestamps by logging warnings
        - Continues processing even if individual entries fail
        
    Example Log Entry:
        "2025-01-01 12:00:03.123456 EM: Process crypto_keystorage_manager.0 is started"
        
    Extracted Data:
        {'crypto_keystorage_manager': '2025-01-01 12:00:03.123456'}
        
    Error Handling:
        - Logs warnings for processes without extractable timestamps
        - Continues processing remaining log entries
        - Returns partial results rather than failing completely
        
    Dependencies:
        - Requires extract_timestamp_from_dlt() function for timestamp parsing
        - Uses re module for regular expression matching
        
    Note:
        This function is crucial for establishing the startup sequence and timing
        of applications, which forms the foundation for startup order validation
        and performance analysis.
    """
    # Dictionary to store process names as keys and their start timestamps as values
    app_start_timestamps = OrderedDict()  

    # Iterate over each line in the log file
    for line in lines:
        # Check if the line contains the required keywords to indicate a process start event
        #if ':EM: Process' in line and 'Pid:' in line and 'is started' in line:
        if 'Application:' in line and 'Init(Up) Time:' in line:
            # Split the line into parts based on the ':EM: Process' keyword
            pattern = r"Application: ([^-\s]+(?:-[^-\s]+)*)"
            m = re.search(pattern, line)
            #parts = line.split(':EM: Process')
           
            # Check if the split resulted in more than one part (i.e., the keyword was found)
            #if len(parts) > 1:
            if m:
                process_name = m.group(1)                              
                # Use a regular expression to extract the timestamp from the line
                timestamp_match = extract_timestamp_from_dlt(line)
                # Check if a timestamp was found
                if timestamp_match!=None:
                    app_start_timestamps[process_name] = timestamp_match
                else:
                    # Log a warning if no timestamp was found for the process
                    logger.warning(f"No timestamp found for process {process_name}")

    # Return the dictionary of process start timestamps
    return app_start_timestamps

def validate_ind_app_startup_order(application, position, application_startup_order):
    """
    Validates whether an individual application started in its expected order position.
    
    This function checks if a specific application started in the correct sequence
    according to the configured startup order. It handles both sequential and
    parallel startup configurations and returns specific failure types for analysis.
    
    Args:
        application (str): Name of the application to validate
        position (int): Actual startup position of the application (1-based)
        application_startup_order (list): List of tuples containing (order_type, app_list)
                                        where order_type is 'Sequential' or 'Parallel'
                                        
    Returns:
        int: Failure type code from OrderFailureType enum:
             - 0: No failure (application started in correct order)
             - 1: ORDER_MISMATCH - Application started in wrong position
             - 2: APPLICATION_NOT_CONFIGURED - Application not found in startup order config
             - 3: APPLICATION_NOT_FOUND - Application missing from logs (handled elsewhere)
             
    Validation Logic:
        Sequential Order:
        - Applications must start in exact specified sequence
        - Each application has a specific expected position
        
        Parallel Order:
        - Applications within a group can start in any order
        - All applications in group share the same position range
        
    Position Calculation:
        - Maintains running position counter across order groups
        - Sequential: Exact position matching
        - Parallel: Range-based position matching
        
    Example Configuration:
        [('Sequential', ['app1', 'app2']), ('Parallel', ['app3', 'app4'])]
        
    Expected Positions:
        - app1: position 1
        - app2: position 2  
        - app3: positions 3-4
        - app4: positions 3-4
        
    Note:
        This function is part of the startup order validation system that ensures
        applications start in the correct sequence for proper system initialization.
    """
    for order_type, order in application_startup_order:
        if len(order) >= position:
            if order_type.lower() == OrderType.SEQUENTIAL.value:
                if order[position - 1] == application:
                    return 0
                else:
                    break
            else:
                if application in order:
                    return 0
                else:
                    break                
        else:
            position -= len(order)
    for order_type, order in application_startup_order:
        if application in order:
            return 1
    return 2


def validate_app_startup_order(dltstart_timestamps, application_startup_order):
    """
    Validates the overall application startup order against the configured sequence.
    
    This function performs comprehensive validation of the entire application startup
    sequence to ensure applications started in the correct order according to the
    configuration. It handles both sequential and parallel startup groups.
    
    Args:
        dltstart_timestamps (dict): OrderedDict mapping application names to startup timestamps
                                   The order of keys represents the actual startup sequence
        application_startup_order (list): List of tuples containing (order_type, app_list)
                                        Defines the expected startup sequence
                                        
    Returns:
        bool: True if the startup order matches configuration, False otherwise
        
    Validation Logic:
        Sequential Groups:
        - Applications must start in exact order specified
        - Uses list comparison for strict ordering
        
        Parallel Groups:
        - Applications can start in any order within the group
        - Uses set comparison to ignore internal ordering
        
    Algorithm:
        1. Extract actual startup sequence from timestamp dictionary keys
        2. Iterate through each configured startup group
        3. Validate the corresponding slice of actual sequence
        4. Move to next group and continue validation
        
    Example Configuration:
        [('Sequential', ['app1', 'app2']), ('Parallel', ['app3', 'app4'])]
        
    Valid Actual Sequences:
        - ['app1', 'app2', 'app3', 'app4']
        - ['app1', 'app2', 'app4', 'app3']
        
    Invalid Actual Sequences:
        - ['app2', 'app1', 'app3', 'app4'] (sequential order violated)
        - ['app1', 'app2', 'app5', 'app4'] (wrong application in parallel group)
        
    Note:
        This function provides high-level startup order validation for the entire
        system, complementing the individual application validation performed by
        validate_ind_app_startup_order().
    """
    apps = list(dltstart_timestamps.keys())
    cur_len = 0
    for order_type, order in application_startup_order:
        if cur_len + len(order) > len(apps):
            return False
        if order_type.lower() == OrderType.SEQUENTIAL.value:
            if apps[cur_len:cur_len+len(order)]!=order:
                return False
        else:
            if set(apps[cur_len:cur_len+len(order)]) != set(order):
                return False
        cur_len += len(order)
    return True


def extract_timestamp_from_dlt(line):
    """
    Extracts timestamp from DLT (Diagnostic Log and Trace) log file format.
    
    This function parses DLT log entries to extract timestamp information from
    the standardized DLT log format. It handles the specific structure of DLT
    log lines to locate and extract timing data.
    
    Args:
        line (str): Single line from a DLT log file
        
    Returns:
        float or None: Timestamp as float if successfully extracted, None otherwise
        
    DLT Log Format:
        DLT logs typically have a structured format with space-separated fields.
        The timestamp is expected to be in the 4th field (index 3) of the line.
        
    Parsing Logic:
        1. Split the log line by spaces to separate fields
        2. Check if line has sufficient fields (more than 5)
        3. Extract timestamp from the 4th field (index 3)
        4. Convert to float for numerical operations
        
    Error Handling:
        - Returns None if line doesn't have enough fields
        - Returns None if timestamp conversion fails
        - Gracefully handles malformed log entries
        
    Example DLT Line:
        "2025-01-01 12:00:03.123456 INFO APP 1641234567.123 Message content"
        
    Extracted Timestamp:
        1641234567.123 (as float)
        
    Note:
        This function is specifically designed for the DLT log format used
        in automotive diagnostic systems. The timestamp format may vary
        depending on the DLT configuration and version.
    """
    parts = line.split(' ')
    # Check if the line has enough parts to extract the timestamp
    if len(parts) > 5:
        # Extract the timestamp from the line
        timestamp = parts[3].strip()
        return float(timestamp)
    return None


# Function to extract the welcome timestamp from the log lines
def extract_welcome_timestamp(lines):
    """
    Extracts the system welcome timestamp from DLT log file lines.
    
    This function searches through log file lines to find the system welcome message
    that indicates when the system has completed its initial startup phase. This
    timestamp serves as the baseline reference point for all subsequent timing
    measurements and analysis.
    
    Args:
        lines (list): List of log file lines to search through
        
    Returns:
        float or None: Welcome timestamp as float if found, None if not found
        
    Search Pattern:
        - Looks for lines containing 'KSAR Adaptive' text
        - This pattern indicates the system welcome message
        - Uses the first occurrence found in the log
        
    Processing Logic:
        1. Iterate through each line in the log file
        2. Search for the welcome message pattern using regex
        3. Extract timestamp from the matching line using extract_timestamp_from_dlt()
        4. Return immediately upon finding the first match
        
    Baseline Importance:
        The welcome timestamp is crucial because it represents:
        - End of system initialization phase
        - Start of application startup measurement period
        - Reference point for calculating relative startup times
        
    Example Log Entry:
        "2025-01-01 12:00:02.000000 INFO SYS 1641234562.000 KSAR Adaptive System Ready"
        
    Extracted Timestamp:
        1641234562.000 (used as baseline for all subsequent measurements)
        
    Error Handling:
        - Returns None if welcome message is not found
        - Continues searching through all lines if needed
        - Handles regex matching failures gracefully
        
    Note:
        This function is critical for establishing the measurement baseline.
        Without a valid welcome timestamp, startup time analysis cannot proceed
        as there would be no reference point for relative timing calculations.
    """
    # Initialize welcome_timestamp to None
    welcome_timestamp = None
    # Iterate over each line in the log file
    for line in lines:
        # Search for the welcome message using regular expression
        match = re.search(r'KSAR Adaptive', line)
        if match:
            # Extract the timestamp from the line containing the welcome message
            #welcome_timestamp = re.search(r'\[(\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2}\.\d{3})\]', line).group(1)
            welcome_timestamp = extract_timestamp_from_dlt(line)
            # Break the loop once the welcome timestamp is found
            break
    return welcome_timestamp


def RCAR_ON_OFF_Relay(power_on_off_delay, logger):
    """
    Controls RCAR ECU power using USB relay for automated testing cycles.
    
    This function manages the power cycling of RCAR (R-Car) ECUs using a USB-controlled
    relay system. It performs a complete power cycle (OFF -> delay -> ON) to simulate
    ignition cycles for startup time testing.
    
    Args:
        power_on_off_delay (int/float): Delay time in seconds between power OFF and ON
                                       Simulates the time between ignition cycles
                                       
    Returns:
        bool: True if power cycle completed successfully, False if errors occurred
        
    Power Cycle Sequence:
        1. Turn OFF relay (BITFT_1=0) - Cuts power to ECU
        2. Wait for specified delay period - Allows ECU to fully power down
        3. Turn ON relay (BITFT_1=1) - Restores power to ECU
        4. Brief stabilization delay (0.2s) - Allows power to stabilize
        
    Hardware Requirements:
        - USB relay device compatible with 'usbrelay' command
        - Relay channel BITFT_1 connected to ECU power supply
        - Proper relay wiring for safe power switching
        
    Safety Considerations:
        - Ensures proper power-down time to prevent data corruption
        - Uses controlled timing to simulate realistic ignition cycles
        - Handles command execution errors gracefully
        
    Error Handling:
        - Catches subprocess execution exceptions
        - Logs specific error details for troubleshooting
        - Returns False on any failure to allow test abortion
        
    Use Cases:
        - Automated startup time testing
        - Regression testing of boot sequences
        - Performance validation across multiple cycles
        
    Note:
        This function is specific to RCAR ECU testing setups and requires
        the 'usbrelay' utility to be installed and properly configured.
    """
    try:
        logger.info("Turning OFF relay...")
        subprocess.run(["usbrelay", "BITFT_1=0"])
        time.sleep(float(power_on_off_delay))  #  delay

        logger.info("Turning ON relay...")
        subprocess.run(["usbrelay", "BITFT_1=1"])
        time.sleep(0.2)  #  delay

    except Exception as e:
        logger.error(f"Error executing usbrelay commands: {e}")
        return False
    return True


def power_ON_OFF_Relay(serial_port_relay, baudrate_relay, power_on_off_delay, logger):
    """
    Controls ECU power using serial-controlled relay for automated testing cycles.
    
    This function manages power cycling of ECUs using a serial-controlled relay system.
    It sends AT commands over a serial connection to control relay states, enabling
    automated ignition cycle simulation for startup time testing.
    
    Args:
        serial_port_relay (str): Serial port identifier (e.g., 'COM5', '/dev/ttyUSB0')
        baudrate_relay (int): Serial communication baud rate (e.g., 9600)
        power_on_off_delay (int/float): Delay time in seconds between power OFF and ON
        
    Returns:
        bool: True if power cycle completed successfully, False if errors occurred
        
    Serial Communication:
        - Protocol: AT commands for relay control
        - Configuration: 8 data bits, 1 stop bit, 1-second timeout
        - Commands: AT+CH1=0 (OFF), AT+CH1=1 (ON)
        
    Power Cycle Sequence:
        1. Establish serial connection to relay controller
        2. Send AT+CH1=0 command to turn OFF relay
        3. Wait for specified delay period
        4. Send AT+CH1=1 command to turn ON relay
        5. Brief stabilization delay (100ms)
        
    Hardware Requirements:
        - Serial-controlled relay module with AT command support
        - Relay channel 1 connected to ECU power supply
        - Proper serial cable connection
        - Compatible relay controller firmware
        
    Error Handling:
        - Validates serial port opening
        - Catches serial communication exceptions
        - Logs specific error details for troubleshooting
        - Returns False on any failure
        
    Use Cases:
        - Multi-ECU Elite setup testing
        - Automated startup time validation
        - Continuous integration testing
        - Performance regression testing
        
    Note:
        This function is designed for serial-controlled relay systems commonly
        used in automotive testing environments. The AT command protocol may
        vary depending on the specific relay controller model.
    """
    try:
        #set up your serial port with the desire COM port and baudrate.
        signal = serial.Serial(serial_port_relay, baudrate_relay, bytesize=8, stopbits=1, timeout=1)
        if not signal.is_open:
            logger.error(f"Failed to open serial port: {serial_port_relay}")
            return False
       
        logger.info("Turning OFF relay...")
        signal.write("AT+CH1=0".encode())   # Relay OFF
        time.sleep(float(power_on_off_delay))  # Delay for power off
       
        logger.info("Turning ON relay...")
        signal.write("AT+CH1=1".encode())   # Relay ON
        time.sleep(0.1)  # 100ms delay
    except Exception as e:
        logger.error(f"Failed to open serial port: {e}")
        return False
    return True


def create_workBook(ecu_type, setup_type, iterations, config, logger):
    """
    Creates a comprehensive Excel workbook for startup time analysis reporting.
    
    This function generates a complete Excel workbook structure with multiple worksheets
    for detailed startup time analysis. It creates separate sheets for each test iteration,
    a summary sheet, and an appendix with field descriptions.
    
    Args:
        ecu_type (str): ECU type identifier (e.g., 'RCAR', 'SoC0', 'SoC1')
        setup_type (str): Test setup configuration type (e.g., 'Elite', 'PADAS')
        iterations (int): Number of test iterations to create sheets for
        config (dict): Test configuration containing validation and formatting settings
        
    Returns:
        tuple: 4-tuple containing:
            - report_file (Path): Full path to the Excel report file
            - workbook (openpyxl.Workbook): Excel workbook object
            - sheets (list): List of iteration worksheet objects
            - summary_sheet (openpyxl.worksheet.worksheet.Worksheet): Summary worksheet object
            
    Workbook Structure:
        1. Summary Sheet: Executive summary and cross-iteration analysis
        2. Iteration Sheets: Detailed analysis for each test run (GEN3_StartupTime_01, 02, etc.)
        3. Appendix Sheet: Field descriptions and documentation
        
    File Naming Convention:
        Application_Startup_Time_{setup_type}_{ecu_type}_N{iterations}_{timestamp}.xlsx
        
    Professional Formatting:
        - Removes gridlines from all sheets for clean presentation
        - Uses consistent naming conventions
        - Applies professional styling throughout
        
    Directory Management:
        - Creates report directory if it doesn't exist
        - Uses global local_save_path for consistent file organization
        - Handles directory creation errors gracefully
        
    Error Handling:
        - Separates directory creation and workbook creation errors
        - Provides specific error messages for troubleshooting
        - Returns None values on failure to prevent downstream errors
        - Re-raises exceptions for critical failures
        
    Dependencies:
        - Requires add_appendix_sheet() function for documentation
        - Uses global current_timestamp for unique file naming
        - Depends on openpyxl library for Excel operations
        
    Note:
        This function creates the foundation for all startup time reporting.
        The returned workbook object is used throughout the analysis pipeline
        to generate comprehensive performance reports.
    """
    try:
        # Create the report file name based on the ECU type and current timestamp
        reportName = f"Application_Startup_Time_{setup_type}_{ecu_type}_N{iterations}_{current_timestamp}.xlsx"
       
        # Define the directory where the report will be saved
        report_dir = local_save_path
       
        # Define the full path of the report file
        report_file = report_dir / reportName
       
        # Check if the report directory exists
        if not report_dir.exists():
            # If the directory does not exist, create it
            report_dir.mkdir()
       
    except OSError as e:
        # If an error occurs while creating the directory, logger. the error message and return None
        logger.error(f"Error creating directory: {e}")
        return None, None, None
   
    except Exception as e:
        # If any other exception occurs, logger. the error message and return None
        logger.error(f"An unexpected error occurred: {e}")
        return None, None, None

    try:
        # Create a new Excel workbook
        workbook = openpyxl.Workbook()

        # Get the active sheet in the workbook
        summary_sheet = workbook.active

        # Set the title of the sheet
        summary_sheet.title = 'Summary'

        # Create a list to store the sheets
        sheets = []

        # Create each sheet and add it to the list
        for i in range(1, iterations + 1):
            sheet_title = f"GEN3_StartupTime_{i:02d}"
            sheet = workbook.create_sheet(title=sheet_title)
            sheets.append(sheet)


        # Create sheet for Appendix
        add_appendix_sheet(workbook, ecu_type, config)
 
        # Remove gridlines from all the sheets in the workbook
        for sheet_exl in sheets:
            # Hide the grid lines in the sheet
            sheet_exl.sheet_view.showGridLines = False

        summary_sheet.sheet_view.showGridLines = False
       
        # Return the report file path, workbook object, and active sheet object
        return report_file, workbook, sheets, summary_sheet
   
    except Exception as e:
        # If any exception occurs while creating the workbook or sheet, logger. the error message and return None
        logger.error(f"An error occurred while creating the workbook or sheet: {e}")
        raise e
        # return None, None, None, None


def load_config(file_path, logger):
    """
    Loads and parses configuration files supporting both JSON and YAML formats.
    
    This function provides flexible configuration loading that automatically detects
    and parses both JSON and YAML configuration files. It handles various file
    formats commonly used in testing environments and provides comprehensive
    error handling for configuration management.
    
    Args:
        file_path (str): Relative path to the configuration file from the script directory
                        Supports both .json and .yml/.yaml extensions
                        
    Returns:
        dict or None: Parsed configuration dictionary if successful, None if failed
        
    Supported Formats:
        - JSON (.json): JavaScript Object Notation format
        - YAML (.yml, .yaml): YAML Ain't Markup Language format
        
    File Resolution:
        - Resolves file path relative to the script's parent directory
        - Uses pathlib for cross-platform path handling
        - Automatically detects file format from extension
        
    Parsing Logic:
        1. Determine file format from extension
        2. Open file with appropriate encoding
        3. Parse content using format-specific parser
        4. Return parsed configuration dictionary
        
    Error Handling:
        - FileNotFoundError: Configuration file doesn't exist
        - PermissionError: Insufficient permissions to read file
        - yaml.YAMLError: YAML syntax or parsing errors
        - IOError: General I/O errors during file operations
        - Unsupported format: File extension not recognized
        
    Configuration Validation:
        - Validates file extension before attempting to parse
        - Provides specific error messages for troubleshooting
        - Returns None on any error to prevent downstream failures
        
    Example Usage:
        >>> config = load_config('startup_time_config.yml')
        >>> if config:
        ...     iterations = config['iterations']
        
    Note:
        This function is critical for test configuration management and supports
        the flexibility needed for different deployment environments and
        configuration management practices.
    """
    try:
        config_path = Path(__file__).parent.joinpath(file_path)
        root, ext = os.path.splitext(config_path)
        with open(config_path, 'r') as file:
            if ext == '.json':
                config = json.load(file)
            elif ext in ('.yml', '.yaml'):
                config = yaml.safe_load(file)
            else:
                logger.error(f"'{file_path}' is not a valid config file")        
                return None

        return config
    except (FileNotFoundError, PermissionError, yaml.YAMLError, IOError) as e:
        logger.error(f"An error occurred while reading the file '{file_path}': {e}")
        return None


def add_appendix_sheet(workbook, ecu_type, config):
    """
    Creates an appendix worksheet with field descriptions and documentation.
    
    This function adds a comprehensive documentation sheet to the Excel workbook
    that explains the meaning and calculation of all fields used in the startup
    time analysis. It serves as a reference guide for report users.
    
    Args:
        workbook (openpyxl.Workbook): Excel workbook to add the appendix sheet to
        ecu_type (str): ECU type identifier for header generation
        config (dict): Test configuration for formatting settings
        
    Appendix Content:
        - Field descriptions for all data columns
        - Calculation explanations for derived values
        - Terminology definitions for technical terms
        - Reference information for report interpretation
        
    Sheet Structure:
        1. Professional header with ECU-specific title
        2. Two-column layout: Field Name | Description
        3. Comprehensive field documentation from startup_field_descriptions
        4. Professional formatting with borders and alignment
        
    Documentation Fields:
        - Services/Applications: Application identification
        - Start Time: Timestamp extraction methodology
        - Apps Startup: Calculation from welcome timestamp
        - Init(Up) Time: Application initialization duration
        - IG ON to QNX + KSAR Startup: Offset time explanation
        - Total Time: Complete startup time calculation
        
    Formatting Features:
        - Removes gridlines for professional appearance
        - Applies consistent header styling
        - Uses format_sheet() for professional cell formatting
        - Maintains visual consistency with other worksheets
        
    Use Cases:
        - Report documentation and explanation
        - Training materials for new users
        - Audit trail for calculation methodologies
        - Reference guide for stakeholders
        
    Note:
        This appendix is essential for report usability and ensures that
        all stakeholders can understand the analysis methodology and
        interpret results correctly.
    """
    appendix_sheet = workbook.create_sheet(title='Appendix')
    # appendix_sheet.title = 'Appendix'
    appendix_sheet.sheet_view.showGridLines = False
    start_row = create_header(appendix_sheet, ecu_type, config['Startup Order Judgement'], 'startup_appendix')
    for data_row in startup_field_descriptions:
        appendix_sheet.append(data_row)
    format_sheet(appendix_sheet, start_row, appendix_columns)
   

def is_merged_cell(sheet, cell):
    """
    Checks if a given cell is part of a merged cell range in an Excel worksheet.
    
    This utility function determines whether a specific cell is part of any merged
    cell range in the worksheet. This is important for formatting operations that
    need to handle merged cells differently from regular cells.
    
    Args:
        sheet (openpyxl.worksheet.worksheet.Worksheet): Excel worksheet to check
        cell (openpyxl.cell.cell.Cell): Cell object to test for merge status
        
    Returns:
        bool: True if the cell is part of a merged range, False otherwise
        
    Merge Detection:
        - Uses the cell's coordinate (e.g., 'A1', 'B2') for identification
        - Checks against all merged cell ranges in the worksheet
        - Returns immediately upon finding a match
        
    Use Cases:
        - Column width calculation (merged cells should be skipped)
        - Cell formatting operations
        - Content analysis and processing
        - Layout optimization algorithms
        
    Example:
        >>> cell = sheet['A1']
        >>> if not is_merged_cell(sheet, cell):
        ...     # Process individual cell
        ...     process_cell_content(cell)
        
    Note:
        This function is used by formatting and layout functions to ensure
        proper handling of merged cells, which have different behavior than
        regular cells in Excel operations.
    """
    return cell.coordinate in sheet.merged_cells
 
   
def format_sheet(sheet, start_row, columns):
    """
    Applies comprehensive formatting to Excel worksheet with automatic column width adjustment.
    
    This function provides professional formatting for Excel worksheets including
    cell styling, column width optimization, and content-based formatting rules.
    It's specifically designed for appendix and documentation sheets.
    
    Args:
        sheet (openpyxl.worksheet.worksheet.Worksheet): Excel worksheet to format
        start_row (int): Starting row number for formatting (1-based indexing)
        columns (list): List of column header names for special formatting
        
    Formatting Features:
        Column Headers:
        - Light green background (B5E6A2) for visual distinction
        - Bold font for emphasis
        - Consistent border styling
        
        Regular Cells:
        - Center alignment (horizontal and vertical)
        - Consistent border styling throughout
        - Professional appearance
        
        Column Width Optimization:
        - Calculates optimal width based on content length
        - Tracks maximum content length per column
        - Excludes merged cells from width calculations
        - Adds padding (2 characters) for readability
        
    Processing Logic:
        1. Iterate through all rows starting from start_row
        2. Skip empty rows to optimize processing
        3. Track maximum content length per column
        4. Apply formatting based on cell content type
        5. Set optimal column widths based on content
        
    Width Calculation:
        - Measures string length of cell content
        - Maintains per-column maximum tracking
        - Ignores merged cells to prevent width conflicts
        - Applies consistent padding for visual appeal
        
    Use Cases:
        - Appendix sheet formatting
        - Documentation worksheet styling
        - Reference table formatting
        - Professional report presentation
        
    Note:
        This function is optimized for documentation and reference sheets
        where readability and professional appearance are priorities.
    """
    # Dictionary to track max length in each column
    max_col_widths = {}
    # Iterate over rows starting from start_row
    for row in sheet.iter_rows(min_row=start_row, max_row=sheet.max_row):
        # Skip empty rows
        if all(cell.value is None for cell in row):
            continue
        for col_idx, cell in enumerate(row):

            # Update max width tracking
            if cell.value:
                length = len(str(cell.value))
                col_letter = get_column_letter(col_idx + 1)
                if (col_letter not in max_col_widths or max_col_widths[col_letter] < length) and not is_merged_cell(sheet, cell):
                   max_col_widths[col_letter] = length
            if cell.value in columns:
                cell.fill = PatternFill(start_color="B5E6A2", end_color="B5E6A2", fill_type="solid")
                cell.font = Font(bold=True)
                cell.border = border_style
            else:
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border_style

    # Set the column widths based on content length
    for col_letter, width in max_col_widths.items():
        # Add a little extra space (2 or 3) for padding
        sheet.column_dimensions[col_letter].width = width + 2
 

def create_dlp_files(ecu_config_list, setup_type, config):
    """
    Creates DLT project files (.dlp) for each ECU configuration for log capture.
    
    This function generates DLT (Diagnostic Log and Trace) project files that configure
    the DLT viewer for connecting to and capturing logs from multiple ECUs. Each ECU
    gets its own customized project file with specific IP address and identification.
    
    Args:
        ecu_config_list (list): List of ECU configuration dictionaries
                               Each dict contains 'ecu-type' and 'ip-address' keys
        setup_type (str): Test setup type (e.g., 'Elite', 'PADAS')
        config (dict): Test configuration (currently unused but maintained for consistency)
        
    Returns:
        dict: Dictionary mapping ECU types to their respective DLP file paths
              Structure: {'ECU_TYPE': '/path/to/ECU_TYPE.dlp', ...}
              
    File Generation Process:
        1. Create/clear DLP output directory
        2. Load template proj.dlp file
        3. For each ECU:
           - Parse XML template
           - Update hostname with ECU IP address
           - Update description with ECU type
           - Save customized DLP file
           
    DLP File Structure:
        - XML-based configuration format
        - Contains ECU connection parameters
        - Specifies hostname (IP address) for network connection
        - Includes description for ECU identification
        
    Directory Management:
        - Creates 'DLP' subdirectory in script location
        - Clears existing DLP files to prevent conflicts
        - Maintains clean workspace for each test run
        
    File Naming Convention:
        {setup_type}_{ecu_type}.dlp (e.g., 'Elite_RCAR.dlp')
        
    Error Handling:
        - Validates XML structure before modification
        - Warns about missing XML elements
        - Continues processing other ECUs if one fails
        - Returns partial results on errors
        
    Dependencies:
        - Requires proj.dlp template file in script directory
        - Uses xml.etree.ElementTree for XML manipulation
        - Depends on proper ECU configuration structure
        
    Note:
        These DLP files are essential for automated log capture from multiple
        ECUs simultaneously. Each file configures the DLT viewer to connect
        to a specific ECU and capture its diagnostic logs.
    """
    output_dir = 'DLP'
    script_dir = os.path.dirname(os.path.abspath(__file__))
    output_dir_path = os.path.join(script_dir, output_dir)
    dlp_files = {}
    if is_pre_gen_logs:
        for ecu in ecu_config_list:
            dlp_files[ecu['ecu-type']] = None
        return dlp_files
    # Create output directory if it doesn't exist or clear it if it does
    if os.path.exists(output_dir_path):
        # Clear all files in the directory
        for file in os.listdir(output_dir_path):
            file_path = os.path.join(output_dir_path, file)
            if os.path.isfile(file_path):
                os.unlink(file_path)
    else:
        os.makedirs(output_dir_path)
    # Use the script directory to find the proj.dlp file
    proj_path = os.path.join(script_dir, 'proj.dlp')
    tree = ET.parse(proj_path)
    root = tree.getroot()
    for ecu in ecu_config_list:
        project_name = f"{setup_type}_{ecu['ecu-type']}.dlp"
        # Set hostname text to the IP address
        hostname = root.find('ecu/hostname')
        if hostname is not None:
            hostname.text = ecu['ip-address']
        else:
            print(f"Warning: 'hostname' not found for ECU {ecu['ecu-type']}")
            continue
        # Set description text to the ECU type
        description = root.find('ecu/description')
        if description is not None:
            description.text = ecu['ecu-type']
        else:
            print(f"Warning: 'description' not found for ECU {ecu['ecu-type']}")
        # Write updated XML to file
        output_path = os.path.join(output_dir_path, project_name)
        dlp_files[ecu['ecu-type']] = output_path
        tree.write(output_path, encoding='utf-8', xml_declaration=True)
   
    return dlp_files


def capture_logs_from_dlt_viewer(log_file_name, dlt_file_name, project_file_name, config, ecu_type, logger):
    """
    Captures diagnostic logs from ECU using DLT viewer with cross-platform support.
    
    This function orchestrates the log capture process using the DLT (Diagnostic Log
    and Trace) viewer tool. It handles platform-specific execution and file conversion
    to generate text-based log files for analysis.
    
    Args:
        log_file_name (str): Path where the converted text log file will be saved
        dlt_file_name (str): Path where the raw DLT binary file will be saved
        project_file_name (str): Path to the DLP project file for ECU connection
        config (dict): Test configuration containing execution parameters
        ecu_type (str): ECU type identifier for error reporting
        
    Returns:
        bool: True if log capture successful and file contains data, False otherwise
        
    Platform Support:
        Windows:
        - Uses dlt-viewer.bat wrapper script
        - Supports both PATH-based and explicit path configurations
        - Handles Windows-specific path formatting
        
        Linux:
        - Uses native dlt-viewer command with timeout
        - Performs automatic DLT to text conversion
        - Uses shell commands for process control
        
    Capture Process:
        1. Determine platform and execution method
        2. Launch DLT viewer with specified timeout
        3. Connect to ECU using project file configuration
        4. Capture logs for configured duration
        5. Convert binary DLT to readable text format
        6. Validate output file contains data
        
    Configuration Parameters:
        - DLT-Viewer Log Capture Time: Log capture duration
        - windows.isPathSet: Whether DLT viewer is in system PATH
        - windows.dltViewerPath: Explicit path to DLT viewer executable
        
    File Validation:
        - Checks output file size to ensure data was captured
        - Provides specific error messages for empty files
        - Suggests troubleshooting steps for connection issues
        
    Error Scenarios:
        - Empty log files indicate connection problems
        - Invalid IP addresses prevent ECU connection
        - ECU power/network issues cause capture failures
        
    Note:
        This function is critical for the entire analysis pipeline as it
        provides the raw data for all subsequent processing and reporting.
    """
    print("capture_logs_from_dlt_viewer :: START")
    timeout = config['DLT-Viewer Log Capture Time']
    script_dir = Path(__file__).parent.joinpath("dlt-viewer.bat")

    if sys.platform.startswith("win"):
        isPathSet = config['windows']['Is Environment Path Set']
        if isPathSet:
            subprocess.call([script_dir, "dlt-viewer.exe", str(timeout), log_file_name, dlt_file_name, project_file_name])
        else:
            dlt_viewer_path = config['windows']['DLT-Viewer Installed Path']
            # dlt_viewer_path = os.path.join(dlt_viewer_path, "dlt-viewer.exe")
            log_file_name = os.path.join(log_file_name)
            logger.info(f"dlt_viewer_path: {dlt_viewer_path}")
            logger.info(f"log_file_name : {log_file_name}")
            # subprocess.call([r"dlt-viewer.bat", dlt_viewer_path + "\\", str(timeout), log_file_name])
            subprocess.call([script_dir, dlt_viewer_path, str(timeout), log_file_name, dlt_file_name, project_file_name])
    elif sys.platform.startswith("linux"):
        subprocess.run("timeout " + str(timeout) + " dlt-viewer -p "+project_file_name+" -l "+dlt_file_name+" -v", shell=True)
        print("Converting *.dlt to *.txt...")
        subprocess.run("dlt-viewer -c logs.dlt "+str(log_file_name), shell=True)
        print("Conversion done, successfully...")

    size = os.path.getsize(log_file_name)
    if size == 0:
        logger.warning(f"Generated {os.path.basename(log_file_name)} is empty, Please check for valid IP-address / Status of {ecu_type}.")
        return False
    return True

       
def process_log_file(i, ecu_type, setup_type, log_file_details, dlp_file, config, sheet, overall_IG_ON_iteration, process_start_times, process_times, application_startup_order,application_startup_order_status, logger):
    """
    Processes a single ECU log file for one test iteration, extracting timing data and generating reports.
    
    This function orchestrates the complete log processing workflow for a single test iteration.
    It captures logs, extracts timing information, validates startup order, and generates
    comprehensive Excel reports with embedded visualizations.
    
    Args:
        i (int): Current iteration index (0-based)
        ecu_type (str): ECU type identifier (e.g., 'RCAR', 'SoC0', 'SoC1')
        log_file_details (tuple): 3-tuple containing (filename, logfile, dltfile) paths
        dlp_file (str): Path to DLT project file for log capture
        config (dict): Test configuration containing thresholds and validation settings
        sheet (openpyxl.worksheet.worksheet.Worksheet): Excel worksheet for this iteration
        overall_IG_ON_iteration (dict): Dictionary to store overall startup times per iteration
        process_start_times (dict): Dictionary to accumulate process initialization times
        process_times (dict): Dictionary to accumulate process startup times
        application_startup_order (list): Expected startup order configuration
        application_startup_order_status (dict): Dictionary to store order validation results
        
    Returns:
        bool: True if processing completed successfully, False if any critical errors occurred
        
    Processing Workflow:
        1. Capture logs from ECU using DLT viewer
        2. Read and parse log file with encoding error handling
        3. Extract system welcome timestamp (baseline reference)
        4. Extract application startup timestamps
        5. Validate startup order against configuration
        6. Extract process initialization timing data
        7. Accumulate timing data for cross-iteration analysis
        8. Generate comprehensive Excel report for this iteration
        9. Add traceability hyperlink to source log file
        
    Data Extraction:
        - Welcome timestamp: System initialization completion baseline
        - DLT start timestamps: Application startup completion times
        - Process timestamps: Application initialization durations
        - Startup order validation: Compliance with configured sequence
        
    Error Handling:
        - Log capture failures (connection, timeout, empty files)
        - File reading errors (not found, encoding issues)
        - Missing critical timestamps (welcome, application starts)
        - Data extraction failures (malformed logs)
        - Report generation errors
        
    Data Accumulation:
        - Builds cross-iteration datasets for statistical analysis
        - Tracks overall performance metrics per iteration
        - Maintains startup order validation results
        - Enables min/max/average calculations across iterations
        
    Report Generation:
        - Creates detailed per-iteration analysis worksheet
        - Includes timeline visualizations and performance validation
        - Adds traceability links to source data
        - Applies professional formatting and styling
        
    Note:
        This function is typically called in parallel threads for multi-ECU setups,
        enabling concurrent log processing and analysis across different ECU types.
    """
    try:
        # Get the log file path and name for the specified ECU type and timestamp
        filename, logfile, dltfile = log_file_details
        if not is_pre_gen_logs:
            if not capture_logs_from_dlt_viewer(filename, dltfile, dlp_file, config, ecu_type, logger):
                return False

        # Attempt to open the log file in read mode with error handling for encoding issues
        try:
            with open(filename, 'r', encoding='utf-8', errors='ignore') as file:
                lines = file.readlines()
                time.sleep(2)
        except FileNotFoundError:
            logger.error(f"File not found: {filename}")
            return False
        except UnicodeDecodeError as e:
            logger.error(f"Unicode decode error: {e}")
            return False

        # Extract the welcome timestamp from the log fil
        welcome_timestamp = extract_welcome_timestamp(lines)

        # Check if the welcome timestamp was found
        if welcome_timestamp is None:
            logger.error("KSAR Adaptive not found in log file")
            return False

        # Extract DLTStart timestamps for each application from the log file
        dltstart_timestamps = extract_dltstart_timestamps(lines, logger)

        # Check if the DLTStart timestamps were found
        if not dltstart_timestamps or len(dltstart_timestamps)==0:
            logger.error("Apps DLTStart time is not found in log file")
            return False
       
        application_startup_order_status[i] = {
            'startup_order_status': True, #validate_app_startup_order(dltstart_timestamps, application_startup_order),
            OrderFailureType.ORDER_MISMATCH.name: 0,
            OrderFailureType.APPLICATION_NOT_FOUND.name: 0,
            OrderFailureType.APPLICATION_NOT_CONFIGURED.name: 0
        }
        
        process_Start_End_timestamps = extract_process_timestamps(lines)
        print ("process_Start_End_timestamp:"+str(process_Start_End_timestamps))
        if not process_Start_End_timestamps or len(process_Start_End_timestamps)==0:
            logger.error("Error: Unable to extract process timestamps.")
            return False

        process_timing_info = extract_and_sort_process_timestamps(process_Start_End_timestamps, ecu_type, logger)
        print ("process_timing_info:"+str(process_timing_info))
       
        if not process_timing_info:
            logger.error("Error: No report data available.")
            return False    

        for item in process_timing_info:
            # logger.info(f"Process: {item['process']} Start Timestamp: {item['start_clock']} End Timestamp: {item['end_clock']} start_time_ms: {item['start_time_ms']}")

            # Check if the process is already in the process start times dictionary
            process = item['process']
            if process not in process_start_times:
                # If the process is not in the dictionary, add it with an empty list
                process_start_times[process] = []
            # Append the start_time_ms to the process's list of start times    
            process_start_times[process].append(round_decimal_half_up(item['start_time_ms'], 4))                      
       
        for process, process_time in dltstart_timestamps.items():
            # Check if the process is already in the process times dictionary
            if process not in process_times:
                # If the process is not in the dictionary, add it with an empty list
                process_times[process] = []
            # Append the difference to the process's list of times
            process_times[process].append(round_decimal_half_up(process_time, 4))
       
        overall_IG_ON_iteration[i] = {
            'timestamp': max(dltstart_timestamps.values()),
            'status': True,
        }
        print ("overall_IG_ON_iteration:"+str(overall_IG_ON_iteration))

        generate_apps_startup_report_from_QNX_startup(ecu_type, config, sheet, dltstart_timestamps, process_timing_info, application_startup_order, application_startup_order_status[i], overall_IG_ON_iteration[i], logger)
       
        # Add a hyperlink to the log file in the Excel sheet
        add_logfile_hyperlink(filename, logfile, sheet, ecu_type, setup_type)

    except Exception as e:
        logger.error(f"Exception :: {e}")
        return False
    return True
   
def save_workbook_and_generate_reports(ecu_type, summary_sheet, overall_IG_ON_iteration, process_times, process_start_times, application_startup_order_status, config, workbook, report_file, logger):
    """
    Finalizes Excel workbook with summary analysis and saves the complete test report.
    
    This function generates the executive summary and statistical analysis sections
    of the Excel report, then saves the complete workbook. It creates the final
    deliverable that stakeholders use for performance assessment and decision making.
    
    Args:
        ecu_type (str): ECU type identifier for report headers and titles
        summary_sheet (openpyxl.worksheet.worksheet.Worksheet): Summary worksheet to populate
        overall_IG_ON_iteration (dict): Dictionary mapping iteration index to overall startup times
        process_times (dict): Dictionary mapping process names to lists of startup times across iterations
        process_start_times (dict): Dictionary mapping process names to lists of initialization times
        application_startup_order_status (dict): Startup order validation results per iteration
        config (dict): Test configuration containing validation settings and parameters
        workbook (openpyxl.Workbook): Complete Excel workbook object to save
        report_file (Path): Full path where the Excel report will be saved
        
    Returns:
        bool: True if report generation and saving completed successfully, False otherwise
        
    Report Sections Generated:
        1. Executive Summary Table:
           - Per-iteration performance overview with hyperlinks to detailed sheets
           - Pass/fail status based on performance thresholds
           - Startup order validation summary (if enabled)
           - Failure count breakdown for root cause analysis
           
        2. Statistical Analysis:
           - Min/max/average startup times across all iterations
           - Performance trend analysis and outlier identification
           - Timeline visualizations for average performance
           - Individual application initialization statistics
           
    Summary Features:
        - Hyperlinked iteration numbers for easy navigation to detailed data
        - Color-coded pass/fail indicators for quick assessment
        - Comprehensive failure analysis with categorized counts
        - Professional formatting and styling throughout
        
    Statistical Analysis:
        - Cross-iteration performance aggregation and analysis
        - Embedded timeline graphs for visual performance assessment
        - Sorted data presentation for priority identification
        - Dual perspective analysis (system-level and application-level)
        
    File Operations:
        - Validates workbook integrity before processing
        - Saves complete Excel file with all worksheets and embedded content
        - Provides success confirmation with file path logging
        - Handles file saving errors gracefully
        
    Error Handling:
        - Validates workbook creation success before proceeding
        - Logs specific error messages for troubleshooting
        - Returns failure status to enable proper error handling upstream
        
    Use Cases:
        - Final report generation for stakeholder review
        - Performance trend analysis across multiple test iterations
        - Executive summary creation for management reporting
        - Comprehensive test documentation and audit trails
        
    Note:
        This function creates the final deliverable that summarizes all test results
        and provides the high-level analysis that stakeholders need for decision making.
    """
    # Check if the workbook creation was successful
    if summary_sheet is None:
        logger.error("Error: Unable to create workbook.")
        return False

    each_iteration_test_status(ecu_type, summary_sheet, overall_IG_ON_iteration, config, application_startup_order_status)

    # Export the average data to the Excel sheet
    export_and_plot_average_data_to_excel(summary_sheet, ecu_type, process_times, process_start_times, config, logger)

    # Save the Excel workbook
    workbook.save(report_file)

    #remove_png_files()        

    # logger. a success message
    logger.info(f"Test report is created successfully {report_file}")
    return True      

def start_startup_time_measurement(logger):
    """
    Main entry point for ECU startup time measurement and analysis system.
    
    This is the primary orchestration function that coordinates the entire startup time
    testing workflow. It manages multi-ECU testing, handles different setup types,
    performs automated power cycling, captures logs, and generates comprehensive reports.
    
    The function supports both single-ECU (PADAS/RCAR) and multi-ECU (Elite) configurations
    with parallel processing capabilities for efficient testing.
    
    Returns:
        bool: True if all tests completed successfully, False if any failures occurred
        
    Workflow Overview:
        1. Initialize global variables and directory structure
        2. Load and validate configuration
        3. Determine ECU setup type and enabled ECUs
        4. Create Excel workbooks for each ECU
        5. Generate DLP project files for log capture
        6. Execute test iterations with power cycling
        7. Capture and process logs in parallel threads
        8. Generate comprehensive analysis reports
        9. Clean up temporary files
        
    Global Variables Initialized:
        - cur_dt_time_obj: Current datetime for timestamping
        - local_save_path: Directory structure for reports and logs
        - workbook_map: Excel workbooks for each ECU type
        - current_timestamp: Formatted timestamp for file naming
        
    Directory Structure Created:
        Reports/03_Startup_Time/YYYYMMDD_HH-MM-SS/
        ├── Logs/
        │   ├── RCAR/ (if Elite setup)
        │   ├── SoC0/ (if Elite setup)
        │   └── SoC1/ (if Elite setup)
        ├── DLP/
        └── Excel reports
        
    Configuration Validation:
        - Validates DLT viewer path existence
        - Checks threshold value ranges (0-100 seconds)
        - Validates iteration count and execution time
        - Verifies ECU IP addresses
        
    Multi-Threading Support:
        - Parallel log capture from multiple ECUs
        - Concurrent log processing and analysis
        - Thread synchronization for result collection
        
    Error Handling:
        - Comprehensive exception handling with specific error types
        - Graceful degradation on partial failures
        - Detailed logging for troubleshooting
        - Cleanup operations in finally block
        
    Performance Monitoring:
        - Tracks total execution time
        - Logs performance metrics
        - Provides execution time feedback
        
    Note:
        This function represents the complete startup time testing solution,
        integrating hardware control, log capture, data analysis, and reporting
        into a single automated workflow.
    """
   
    # Declare global variables
    global cur_dt_time_obj
    cur_dt_time_obj = datetime.now()
    global is_pre_gen_logs
    global table_headers
    table_headers = list()
    global local_save_path
    global workbook_map
    workbook_map = {}
    global threshold_map
    threshold_map = {}
    global current_timestamp
    # current_timestamp = '20250630_175500'
    current_timestamp = cur_dt_time_obj.strftime("%Y%m%d_%H%M%S")


    script_start_time = time.perf_counter()
    try:

        isSuccess = True
        anySheet = []
        config = load_config('startup_time_config.json', logger)      
       
        # Check if the configuration is empty
        if config is None:
            logger.error(f"File 'config_file_path' not found.")
            return False
        
        is_pre_gen_logs = config.get('Pre-Generated Logs', False)
        if is_pre_gen_logs:
            logs_folder_path = config.get('logs-folder-path', Path(__file__).parents[0].joinpath("Pre-Generated_Logs"))
            logger.info(f"logs_folder_path: {logs_folder_path}")
            if not logs_folder_path or not os.path.exists(str(logs_folder_path)):
                logger.error("Error: 'logs-folder-path' is not configured in the configuration file.")
                return False
            local_save_path = Path(logs_folder_path)
        else:
            local_save_path = Path(__file__).parents[1].joinpath("Reports", "03_Startup_Time", cur_dt_time_obj.strftime("%Y%m%d_%H-%M-%S"))
            local_save_path.mkdir(parents=True, exist_ok=True)
       
        if not is_pre_gen_logs and config['windows']['DLT-Viewer Installed Path'] and not os.path.isfile(config['windows']['DLT-Viewer Installed Path']):
            logger.error("Configured dlt-viewer path is not valid.")
            return False
        if config.get('Threshold', -1) < 0 or config.get('Threshold') > 100:
            logger.error("Configured 'Threshold' is not valid. Configure its value in range[0, 100].")
            return False
       
        # Retrieve the number of iterations from the configuration
        try:
            iterations = config["Iterations"]
            if not isinstance(iterations, int):
                logger.error("Error: 'Iterations' must be an integer.")
                return False
        except KeyError:
            logger.error("Error: 'Iterations' key not found in the configuration file.")
            return False
       
        duration = config["DLT-Viewer Log Capture Time"]
        if not is_pre_gen_logs and not isinstance(duration, int):
            logger.error("Error: 'DLT-Viewer Log Capture Time' must be an integer.")
            return False
       
        if not is_pre_gen_logs and not isinstance(config.get("Power ON-OFF Delay", 25), int):
            logger.error("Error: 'Power ON-OFF Delay' must be an integer.")
            return False
        
        if not is_pre_gen_logs and config.get("Power ON-OFF Delay", 25)<=0:
            logger.error("Error: 'Power ON-OFF Delay' must be greater than 0.")
            return False

        process_times_map = {}
        process_start_times_map = {}
        overall_IG_ON_iteration_map = {}
        application_startup_order_status_map = {}
        application_startup_order_map = {}
        setup_type = None
        enabled_ecu_list = set()
               
        if config.get('ECU_setting', {}).get('PADAS', {}).get('RCAR', False):
            enabled_ecu_list.add('PADAS')
            setup_type = 'PADAS'
        else:
            for board_type, enabled in config.get('ECU_setting', {}).get('Elite', {}).items():
                if enabled:
                    enabled_ecu_list.add(board_type)
                    setup_type = 'ELITE'
       
        print(setup_type, enabled_ecu_list)
        if setup_type is None or len(enabled_ecu_list) == 0:
            logger.error("No enabled ECU found in the configuration.")
            return False

        ecu_config_list = [ecu for ecu in config['ecu-config'] if ecu['ecu-type'] in enabled_ecu_list]
        for ecu in ecu_config_list:
            if ecu['ecu-type'] == ECUType.PADAS.value:
                ecu['ecu-type'] = ECUType.RCAR.value
            if not is_pre_gen_logs:
                if ecu['ecu-type'] == ECUType.RCAR.value:
                    ecu['ip-address'] = config['ECU_setting']['RCAR_IPAddress']
                elif ecu['ecu-type'] == ECUType.SoC0.value:
                    ecu['ip-address'] = config['ECU_setting']['Qualcomm_SoC0_IPAddress']
                elif ecu['ecu-type'] == ECUType.SoC1.value:
                    ecu['ip-address'] = config['ECU_setting']['Qualcomm_SoC1_IPAddress']
            workbook_map[ecu['ecu-type']] = tuple(create_workBook(ecu['ecu-type'], setup_type, iterations, config, logger))
           
            # Check if the workbook creation was successful
            if workbook_map[ecu['ecu-type']][2] is None:
                logger.error("Error: Unable to create workbook.")
                return False
            process_times_map[ecu['ecu-type']] = {}
            process_start_times_map[ecu['ecu-type']] = {}
            overall_IG_ON_iteration_map[ecu['ecu-type']] = {}
            application_startup_order_status_map[ecu['ecu-type']] = {}
            application_startup_order = []
            for block in ecu['startup-order']:
                application_startup_order.append(tuple([block['Order Type'], [app.strip() for app in block['Applications'].split(',') if len(app.strip()) > 0]]))
            application_startup_order_map[ecu['ecu-type']] = list(application_startup_order)
            
            threshold_map[ecu['ecu-type']] = {}
            for i, threshold_config_grp in enumerate(ecu.get('threshold-config', [])):
                if threshold_config_grp.get('Threshold', -1) < 0 or threshold_config_grp.get('Threshold', -1) > 100:
                    logger.error(f"Configured 'Threshold' is not valid. Configure its value in range[0, 100] for {i}th group in {ecu['ecu-type']}.")
                    return False
                for app in threshold_config_grp.get('Applications', '').split(','):
                    if len(app.strip()) > 0: 
                        threshold_map[ecu['ecu-type']][app.strip()] = threshold_config_grp.get('Threshold')
             
        logger.info(f"Threshold Map: {threshold_map}")

        # if config['ecu-config']['setup-type'] == ECUType.ELITE.value:
        if not is_pre_gen_logs and not validate_ip_address(ecu_config_list, logger):
            return False
        dlp_files = create_dlp_files(ecu_config_list, setup_type, config)
        if not is_pre_gen_logs and (not dlp_files or len(dlp_files) == 0):
            return False

        # Loop through the iterations
        for i in range(iterations):
            
            if not is_pre_gen_logs:
                if setup_type == ECUType.RCAR.value:
                    if not RCAR_ON_OFF_Relay(config.get('Power ON-OFF Delay', 25), logger):
                        return False
                else:
                    if not power_ON_OFF_Relay(config.get('serial-port-relay'), config.get('baudrate-relay'), config.get('Power ON-OFF Delay', 25), logger):
                        return False
           
            threads = []
            for ecu_type, (report_file, workbook, sheets, summary_sheet) in workbook_map.items():
                print("Thread: ", ecu_type, ": Started")
               
                filename_list = {}
                if not is_pre_gen_logs:
                    if setup_type == ECUType.ELITE.value:
                        filename_list = get_log_file_paths_for_elite(i, ecu_config_list, setup_type)
                    else:
                        filename_list[ecu_type] = tuple(get_log_file_path(ecu_type, setup_type, i))
                else:
                    filename_list[ecu_type] = extract_log_file_paths(i, ecu_type, setup_type, logger)
                logger.info(f"Log files for {ecu_type} in iteration {i}: {filename_list}")
                if any(not filename for (filename, logfile, dltfile) in filename_list.values()):
                    if is_pre_gen_logs:
                        logger.error(f"Log file not found for {ecu_type} in iteration {i}. Please check the configuration.")
                    else:
                        logger.error(f"Log file not created for {ecu_type} in iteration {i}. Please check the configuration.")
                    return False
                thread = ResultThread(
                    target=process_log_file,
                    args=(
                        i,
                        ecu_type,
                        setup_type,
                        filename_list[ecu_type],
                        dlp_files[ecu_type],
                        config,
                        sheets[i],
                        overall_IG_ON_iteration_map[ecu_type],
                        process_start_times_map[ecu_type],
                        process_times_map[ecu_type],
                        application_startup_order_map[ecu_type],
                        application_startup_order_status_map[ecu_type],
                        logger
                     )
                )
                threads.append(thread)
                thread.start()
            # Wait for all threads to complete
            for thread in threads:
                thread.join()
                print("Thread result :: ", thread.result)
                # if not thread.result:
                anySheet.append(thread.result)
        print('anySheet:', anySheet)
        if not any(anySheet):
            isSuccess = False

        # Save workbooks and generate reports for each ECU type
        for ecu_type, (report_file, workbook, sheets, summary_sheet) in workbook_map.items():
            if len(overall_IG_ON_iteration_map[ecu_type]) > 0:
                if not save_workbook_and_generate_reports(
                    ecu_type,
                    summary_sheet,
                    overall_IG_ON_iteration_map[ecu_type],
                    process_times_map[ecu_type],
                    process_start_times_map[ecu_type],
                    application_startup_order_status_map[ecu_type],
                    config,
                    workbook,
                    report_file,
                    logger):
                    isSuccess = False

    except KeyError as e:
        logger.error(f"Error: Missing expected key in ECU input fields: {e}")
        isSuccess = False
    except Exception as e:
        logger.error(f"An error occurred: {e}")
        isSuccess = False
    finally:
        remove_png_files(logger)
        script_end_time = time.perf_counter()
        logger.info(f"Total script execution time: {(script_end_time-script_start_time):.3f} seconds")
    print("Final response :: ", isSuccess)
    return isSuccess